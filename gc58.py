#!/usr/bin/env python

# -*- coding: utf-8 -*-


#GC58

#Requirements:
#openpyxl 2.3.3.
#jdcal 1.2
#et_xmlfile-1.0.1
#Python 2.7 (2.7.10)

#Licenses

GCLH = "GlossaryCheck LICENSE"
GCLT= "GlossaryCheck is written by A.D.Klumpp using Python and the Python library openpyxl including jdcal and et_xmlfile (see license texts below or in the folders of the libraries). GlossaryCheck is released under the terms of the GNU General Public License. Copyright (C) 2015 A.D.Klumpp. GlossaryCheck is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY. The full copyright notices and the full license texts shall be included in all copies or substantial portions of the Software."

OPLH = "OPENPYXL 2.3.3 LICENSE"
OPLT = "(http://openpyxl.readthedocs.org/en/latest/_modules/openpyxl/worksheet/header_footer.html) Copyright (c) 2010-2015 openpyxl. Permission is hereby granted, free of charge, to any person obtaining a copy of this software and associated documentation files (the 'Software'), to deal in the Software without restriction, including without limitation the rights to use, copy, modify, merge, publish, distribute, sublicense, and/or sell copies of the Software, and to permit persons to whom the Software is furnished to do so, subject to the following conditions: The above copyright notice and this permission notice shall be included in all copies or substantial portions of the Software. THE SOFTWARE IS PROVIDED 'AS IS', WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE."

JDCALH = "jdcal 1.2 LICENSE"
JDCALT = '(https://pypi.python.org/pypi/jdcal) Copyright (c) 2011, Prasanth Nair. All rights reserved.\nRedistribution and use in source and binary forms, with or without modification, are permitted provided that the following conditions are met:\n1. Redistributions of source code must retain the above copyright notice, this list of conditions and the following disclaimer.\n2. Redistributions in binary form must reproduce the above copyright notice, this list of conditions and the following disclaimer in the documentation and/or other materials provided with the distribution.\nTHIS SOFTWARE IS PROVIDED BY THE COPYRIGHT HOLDERS AND CONTRIBUTORS "AS IS" AND ANY EXPRESS OR IMPLIED WARRANTIES, INCLUDING, BUT NOT LIMITED TO, THE IMPLIED WARRANTIES OF MERCHANTABILITY AND FITNESS FOR A PARTICULAR PURPOSE ARE DISCLAIMED. IN NO EVENT SHALL THE COPYRIGHT HOLDER OR CONTRIBUTORS BE LIABLE FOR ANY DIRECT, INDIRECT, INCIDENTAL, SPECIAL, EXEMPLARY, OR CONSEQUENTIAL DAMAGES (INCLUDING, BUT NOT LIMITED TO, PROCUREMENT OF SUBSTITUTE GOODS OR SERVICES; LOSS OF USE, DATA, OR PROFITS; OR BUSINESS INTERRUPTION) HOWEVER CAUSED AND ON ANY THEORY OF LIABILITY, WHETHER IN CONTRACT, STRICT LIABILITY, OR TORT (INCLUDING NEGLIGENCE OR OTHERWISE) ARISING IN ANY WAY OUT OF THE USE OF THIS SOFTWARE, EVEN IF ADVISED OF THE POSSIBILITY OF SUCH DAMAGE.'

ETXMLLH = "et_xmlfile"
ETXMLLT = "et_xmlfile is a low memory library for creating large XML files. It is based upon the xmlfile module from lxml <http://lxml.de/api.html#incremental-xml-generation>_ with the aim of allowing code to be developed that will work with both libraries. It was developed initially for the openpyxl project but is now a standalone module. The code was written by Elias Rabel as part of the Python Duesseldorf <http://pyddf.de>_ openpyxl sprint in September 2014. Version: 1.0.1. License: MIT. Home-page: https://bitbucket.org/openpyxl/et_xmlfile"

#KVLH = "KIVY LICENSE"
#KVLT = "(https://github.com/kivy/kivy/blob/master/LICENSE) Copyright (c) 2010-2015 Kivy Team and other contributors. Permission is hereby granted, free of charge, to any person obtaining a copy of this software and associated documentation files (the 'Software'), to deal in the Software without restriction, including without limitation the rights to use, copy, modify, merge, publish, distribute, sublicense, and/or sell copies of the Software, and to permit persons to whom the Software is furnished to do so, subject to the following conditions: The above copyright notice and this permission notice shall be included in all copies or substantial portions of the Software. THE SOFTWARE IS PROVIDED 'AS IS', WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE."

GPLH = "GNU GENERAL PUBLIC LICENSE"
GPLT = "See http://www.gnu.org/licenses/"

PLH = "PYTHON LICENSE"
PLT = "See https://www.python.org/download/releases/2.7.6/license/"


#sheet of input .xlsx
# 0=1, 1=2...

sheetgloss = 0
#sheet glossary

sheetstrng = 1
#sheet strings


firstline = 1

#input:
ceg = 1
# Column 0: Glossary English (eg)

ctg = 2
# Column 1: Glossary Translation (tg)

csid = 1
#csid = 2
# Colunm 2: String ID (sid)

ces = 2
# Column 3: String English (es)

cts = 3
# Column 4: String Translation (ts)


  
           

           
            
         
            



#import xlrd
import re
#from xlrd import open_workbook,cellname, XL_CELL_TEXT, cellnameabs, colname
from Tkinter import *
import tkFileDialog
import Tkinter, tkSimpleDialog
from tkFileDialog import askopenfilename
from Tkinter import Frame, Tk, BOTH, Text, Menu, END
import subprocess as sub
import tkMessageBox
import Tkinter
import pickle
import datetime
#import xlsxwriter
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from collections import Counter
import itertools
import sys
reload(sys)
sys.setdefaultencoding('utf-8')




class GCMain(Frame):




   

    def __init__(self, parent):
        Frame.__init__(self, parent)   

       
        self.parent = parent        
        self.initUI()

        frame = Frame(parent)
        frame.pack()

        


        Label(parent, text="GlossaryCheck\n\n\n").pack()
        
   
        Label(parent, text="For further info please read the Manual and the LICENSE-texts (START-Menu).\n\n\n\n").pack()


            


    def initUI(self):

        self.parent.title('GlossaryCheck')
        self.pack(fill=BOTH, expand=1)

        menubar = Menu(self.parent)
        self.parent.config(menu=menubar)

                    

        
        fileMenu = Menu(menubar)
      
       
        menubar.add_cascade(label="START", menu=fileMenu)

       

        fileMenu.add_command(label="GlossaryCheck", command=self.GlossaryCheck)
        fileMenu.add_command(label="GC MatchCase", command=self.GCMatchCase)
        fileMenu.add_command(label="---------------------", command=None)
        fileMenu.add_command(label="GCList", command=self.GCList)
        fileMenu.add_command(label="GCList(Split)", command=self.GCListSplit)
        fileMenu.add_command(label="GCList(Cut)", command=self.GCListCut)
        fileMenu.add_command(label="---------------------", command=None)        
        fileMenu.add_command(label="NumberCheck", command=self.NumberCheck)
        fileMenu.add_command(label="GlossaryCreator", command=self.GCCreator)
        fileMenu.add_command(label="---------------------", command=None)
        fileMenu.add_command(label="Manual", command=self.Manual)
        #fileMenu.add_command(label="InputTemplate", command=self.InputTempl)
        fileMenu.add_command(label="Legal/Licenses", command=self.Legal)



    def InputTempl(self):

            year = datetime.datetime.now().year
            month = datetime.datetime.now().month
            hour = datetime.datetime.now().hour
            minute = datetime.datetime.now().minute
            day = datetime.datetime.now().day
            second = datetime.datetime.now().second
    

     


            templ_filename = 'InputTemplate' + str(day) + str(hour) + str(minute) + str(second) + '.xlsx'


            wbt = Workbook()

            ws1 = wbt.active
            #ws1 = wbt.create_sheet()
            ws1.title = "Glossary"
            #ws1.cell(row=1, column=1).value = 'GID'
            ws1.cell(row=1, column=1).value = 'Source Term'
            ws1.cell(row=1, column=2).value = 'Translation'


            ws2 = wbt.create_sheet()
            ws2.title = "Strings"
            ws2.cell(row=1, column=1).value = 'String ID'
            ws2.cell(row=1, column=2).value = 'Source String'
            ws2.cell(row=1, column=3).value = 'Translation'


            wbt.save(filename = templ_filename)

            
            
            #ws1.column_dimensions["A"].width = 30
           

            master = Tk()

            w = Label(master, text="\n\n    InputTemplate created. The Template file ('InputTemplate... .xlsx') is in the GlossaryCheck folder.    \n\n", bg="green")
            w.pack()

            mainloop()

          


    def GlossaryCheck(self):



  
     
  

        #ftypes = [('Excel files', '.xls'),('Excel XML', '*.xlsx'), ('All files', '*')]
        ftypes = [('Excel files', '.xlsx')]
        dlg = tkFileDialog.Open(self, filetypes = ftypes)
        fl = dlg.show()

        




        if fl != '':

            
            filename = fl
            book = load_workbook(filename)
            sheetgl = book.worksheets[sheetgloss]
            sheetst = book.worksheets[sheetstrng]




            year = datetime.datetime.now().year
            month = datetime.datetime.now().month
            hour = datetime.datetime.now().hour
            minute = datetime.datetime.now().minute
            day = datetime.datetime.now().day
            second = datetime.datetime.now().second
    

            #filename = str (year) + str(month)
            #filename = str(year) + '_' + str(month) + '_' + str(day) + '_' + str(hour) + '_' + str(minute) + str(second) + '.txt'
            #filenamenb = 'NumberIssues' + str(day) + str(hour) + str(minute) + str(second) + '.txt'
            filenameGI = 'GlossaryCheck' + str(day) + str(hour) + str(minute) + str(second) + '.xlsx'
          



            #f = open(filenamenb, 'w')


            
 # Create a  workbook and add a worksheet.
            #workbookxw = xlsxwriter.Workbook(filenamexw)
            #worksheetxw = workbookxw.add_worksheet()
            wbgi = Workbook()

            wsgi1 = wbgi.active
            wsgi1.title = "GlossaryCheck"

 

            #rowxw = 1

            nu=1


                        
                    

            #row_countgl = sheetgl.get_highest_row()+1
            #row_countst = sheetst.get_highest_row()+1


           
            row_countgl = sheetgl.max_row+1
            row_countst = sheetst.max_row+1


            
         
       


            #for m in range(int(firstline), int(row_countgl)):
            for n in range(int(firstline), int(row_countst)):

                   
                    #textpr = "In progress..."
                    #sys.stdout.write(str(textpr))
                    #sys.stdout.flush()

                    #eg = sheetgl.cell(row = m,column = ceg).value
                    #tg = sheetgl.cell(row = m,column = ctg).value

                    sid = sheetst.cell(row = n,column = csid).value
                    es = sheetst.cell(row = n,column = ces).value
                    es = str(es)
                    ts = sheetst.cell(row = n,column = cts).value
                    ts = str(ts)

                   
                    if es != None:                        
                        esl = es.lower()
                        esl = esl.replace('\n', ' ').replace('\r', '')
                    if es == None:
                        esl = "string missing"

                    
                    if ts != None:                        
                        tsl = ts.lower()
                        tsl = tsl.replace('\n', ' ').replace('\r', '')
                    if ts == None:
                        tsl = "string missing"

                 

                    
                    
                                       
                    #for n in range(int(firstline), int(row_countst)):
                    for m in range(int(firstline), int(row_countgl)):

          
                        #sid = sheetst.cell(row = n,column = csid).value
                        #es = sheetst.cell(row = n,column = ces).value
                        #ts = sheetst.cell(row = n,column = cts).value                  

                        eg = sheetgl.cell(row = m,column = ceg).value
                        eg = str(eg)
                        tg = sheetgl.cell(row = m,column = ctg).value
                        tg = str(tg)



                        #remove leading/trailing spaces, case insensitiv

                        if eg != None and "." not in eg:
                            eg = sheetgl.cell(row = m,column = ceg).value.strip()
                            egl = eg.lower()
                        if eg != None and "." in eg:
                            egl = eg
                        if eg == None:
                            eg = "None"
                            egl = "string missing"
                            
                        if tg != None and "." not in tg:  
                            tg = sheetgl.cell(row = m,column = ctg).value.strip()
                            tgl = tg.lower()
                        if tg != None and "." in tg:
                            tgl = eg                         
                        if tg == None:
                            tg = "None"
                            tgl = "string missing"
                            

              
                        

                      
                        #if sheet.cell(m,0).value in sheet.cell(n,2).value and sheet.cell(m,1).value not in sheet.cell(n,3).value:
                        #if eg in es and tg not in ts and ts != "":
                        #if unicode(eg) in unicode(es) and unicode(tg) not in unicode(ts) and unicode(ts) != unicode(""):
                        #if unicode(eg) in unicode(es) and unicode(tg) not in unicode(ts) or unicode(tg) in unicode(ts) and unicode(eg) not in unicode(es):

                        
                       

                        #if unicode(egl) in unicode(esl) and unicode(tgl) not in unicode(tsl):

                        #if re.search(r'\b' + unicode(eg) + r'\b', unicode(es)) and unicode(tg) not in unicode(ts) and unicode(tg)!= "None":



                        #if unicode(egl) in unicode(esl) and unicode(tgl) not in unicode(tsl) and unicode(tg)!= None: 

                        if re.search(r'\b' + unicode(egl) + r'\b', unicode(esl),  re.UNICODE) and unicode(tgl) not in unicode(tsl) and unicode(tg)!= None:




                                   


                                    nu=nu+1
                      
                        



                                    #worksheetxw.write(nu, 1, eg)
                                    #worksheetxw.write(nu, 2, tg)
                                    #worksheetxw.write(nu, 3, sid)
                                    #worksheetxw.write(nu, 4, es)
                                    #worksheetxw.write(nu, 5, ts)


                                    wsgi1.cell(row=nu, column=1).value = eg
                                    wsgi1.cell(row=nu, column=2).value = tg
                                    wsgi1.cell(row=nu, column=3).value = sid
                                    wsgi1.cell(row=nu, column=4).value = es
                                    wsgi1.cell(row=nu, column=5).value = ts



                                    #print eg

                                    #print "In progress... "

                                    sys.stdout.write("\rGlossaryCheck data analysis in progress. This can take several minutes... ")
                                    sys.stdout.flush()


            #worksheetxw.write(1, 1, 'eg')
            #worksheetxw.write(1, 2, 'tg')
            #worksheetxw.write(1, 3, 'sid')
            #worksheetxw.write(1, 4, 'es')
            #worksheetxw.write(1, 5, 'ts')
            
            #worksheetxw.set_default_row(hide_unused_rows=True)
                                
            #workbookxw.close()

            wsgi1.cell(row=1, column=1).value = 'Source Glossary'
            wsgi1.cell(row=1, column=2).value = 'Loc. Glossary'
            wsgi1.cell(row=1, column=3).value = 'String ID'
            wsgi1.cell(row=1, column=4).value = 'Source String'
            wsgi1.cell(row=1, column=5).value = 'Loc. String'

            wbgi.save(filename = filenameGI)

            

            print 'Done'

            

            master = Tk()

            w = Label(master, text="\n\n    The output file ('GlossaryCheck... .xlsx') is in the GlossaryCheck folder.    \n\n", bg="green")
            w.pack()

            mainloop()







    def GCMatchCase(self):


        
  
     
  

        #ftypes = [('Excel files', '.xls'),('Excel XML', '*.xlsx'), ('All files', '*')]
        ftypes = [('Excel files', '.xlsx')]
        dlg = tkFileDialog.Open(self, filetypes = ftypes)
        fl = dlg.show()



        if fl != '':
            filename = fl
            book = load_workbook(filename)
            sheetgl = book.worksheets[sheetgloss]
            sheetst = book.worksheets[sheetstrng]
         


            year = datetime.datetime.now().year
            month = datetime.datetime.now().month
            hour = datetime.datetime.now().hour
            minute = datetime.datetime.now().minute
            day = datetime.datetime.now().day
            second = datetime.datetime.now().second
    

            #filename = str (year) + str(month)
            #filename = str(year) + '_' + str(month) + '_' + str(day) + '_' + str(hour) + '_' + str(minute) + str(second) + '.txt'
            #filenamenb = 'NumberIssues' + str(day) + str(hour) + str(minute) + str(second) + '.txt'
            filenameGI = 'GCMatchCase' + str(day) + str(hour) + str(minute) + str(second) + '.xlsx'
          

            #print filename

            #f = open(filenamenb, 'w')



            


          

             
           # print >> f
            #print >> f, "Terminology error (term/string ID):"
            #print >> f, "Terminology error (term/string ID):"
            #print >> f
            
 # Create a xlsxwriter workbook and add a worksheet.
            #workbookxw = xlsxwriter.Workbook(filenamexw)
            #worksheetxw = workbookxw.add_worksheet()
            wbgi = Workbook()

            wsgi1 = wbgi.active
            wsgi1.title = "GCMatchCase"

 

            #rowxw = 1

            nu=1


                        
                    

            #row_countgl = sheetgl.get_highest_row()+1
            #row_countst = sheetst.get_highest_row()+1



            row_countgl = sheetgl.max_row+1
            row_countst = sheetst.max_row+1            




           


            #for m in range(int(firstline), int(row_countgl)):
            for n in range(int(firstline), int(row_countst)):

                    #eg = sheetgl.cell(row = m,column = ceg).value
                    #tg = sheetgl.cell(row = m,column = ctg).value

                    sid = sheetst.cell(row = n,column = csid).value
                    es = sheetst.cell(row = n,column = ces).value
                    es = str(es)
                    ts = sheetst.cell(row = n,column = cts).value
                    ts = str(ts)

                    #sidl = sheetst.cell(row = n,column = csid).value.lower()
                    #esl = sheetst.cell(row = n,column = ces).value.lower()
                    #tsl = sheetst.cell(row = n,column = cts).value.lower()


                    
                    if es != None:                        
                       
                        es = es.replace('\n', ' ').replace('\r', '')
                    if es == None:
                        es = "string missing"

                    
                    if ts != None:                        
                       
                        ts = ts.replace('\n', ' ').replace('\r', '')
                    if ts == None:
                        ts = "string missing"



                    

                                       
                    #for n in range(int(firstline), int(row_countst)):
                    for m in range(int(firstline), int(row_countgl)):

          
                        #sid = sheetst.cell(row = n,column = csid).value
                        #es = sheetst.cell(row = n,column = ces).value
                        #ts = sheetst.cell(row = n,column = cts).value




                     

                        eg = sheetgl.cell(row = m,column = ceg).value
                        eg = str(eg)
                        tg = sheetgl.cell(row = m,column = ctg).value
                        tg = str(tg)

                        #remove leading/trailing spaces

                        if eg != None:
                            eg = sheetgl.cell(row = m,column = ceg).value.strip()
                        if eg == None:
                            eg = "None"
                        if tg != None:  
                            tg = sheetgl.cell(row = m,column = ctg).value.strip()
                        if tg == None:
                            tg = "None"
                            


                            

                            

                        #egl = sheetgl.cell(row = m,column = ceg).value.lower()
                        #tgl = sheetgl.cell(row = m,column = ctg).value.lower()

                       
                       
                        
                        

                      
                        #if sheet.cell(m,0).value in sheet.cell(n,2).value and sheet.cell(m,1).value not in sheet.cell(n,3).value:
                        #if eg in es and tg not in ts and ts != "":
                        #if unicode(eg) in unicode(es) and unicode(tg) not in unicode(ts) and unicode(ts) != unicode(""):
                        #if unicode(eg) in unicode(es) and unicode(tg) not in unicode(ts) or unicode(tg) in unicode(ts) and unicode(eg) not in unicode(es):

                        
                        #if unicode(egl) in unicode(esl) and unicode(tgl) not in unicode(tsl) and unicode(tgl)!= "None":

                        #if re.search(r'\b' + unicode(eg) + r'\b', unicode(es)) and unicode(tg) not in unicode(ts) and unicode(tg)!= "None":

                        if re.search(r'\b' + unicode(eg) + r'\b', unicode(es),  re.UNICODE) and unicode(tg) not in unicode(ts) and unicode(tg)!= "None":



                


                                    nu=nu+1
                      
                        



                                    #worksheetxw.write(nu, 1, eg)
                                    #worksheetxw.write(nu, 2, tg)
                                    #worksheetxw.write(nu, 3, sid)
                                    #worksheetxw.write(nu, 4, es)
                                    #worksheetxw.write(nu, 5, ts)


                                    wsgi1.cell(row=nu, column=1).value = eg
                                    wsgi1.cell(row=nu, column=2).value = tg
                                    wsgi1.cell(row=nu, column=3).value = sid
                                    wsgi1.cell(row=nu, column=4).value = es
                                    wsgi1.cell(row=nu, column=5).value = ts


                                    sys.stdout.write("\rGCMatchCase data analysis in progress. This can take several minutes... ")
                                    sys.stdout.flush()



                                    ###print eg



            #worksheetxw.write(1, 1, 'eg')
            #worksheetxw.write(1, 2, 'tg')
            #worksheetxw.write(1, 3, 'sid')
            #worksheetxw.write(1, 4, 'es')
            #worksheetxw.write(1, 5, 'ts')
            
            #worksheetxw.set_default_row(hide_unused_rows=True)
                                
            #workbookxw.close()

            wsgi1.cell(row=1, column=1).value = 'Source Glossary'
            wsgi1.cell(row=1, column=2).value = 'Loc. Glossary'
            wsgi1.cell(row=1, column=3).value = 'String ID'
            wsgi1.cell(row=1, column=4).value = 'Source String'
            wsgi1.cell(row=1, column=5).value = 'Loc. String'

            wbgi.save(filename = filenameGI)

            print 'Done'

            #print 'Terminology check done'

            master = Tk()

            w = Label(master, text="\n\n    The output file ('GCMatchCase... .xlsx') is in the GlossaryCheck folder.    \n\n", bg="green")
            w.pack()

            mainloop()
            













    def GCListSplit(self):


        
  
     
  

        #ftypes = [('Excel files', '.xls'),('Excel XML', '*.xlsx'), ('All files', '*')]
        ftypes = [('Excel files', '.xlsx')]
        dlg = tkFileDialog.Open(self, filetypes = ftypes)
        fl = dlg.show()



        if fl != '':
            filename = fl
            book = load_workbook(filename)
            sheetgl = book.worksheets[sheetgloss]
            sheetst = book.worksheets[sheetstrng]
         


            year = datetime.datetime.now().year
            month = datetime.datetime.now().month
            hour = datetime.datetime.now().hour
            minute = datetime.datetime.now().minute
            day = datetime.datetime.now().day
            second = datetime.datetime.now().second
    

            #filename = str (year) + str(month)
            #filename = str(year) + '_' + str(month) + '_' + str(day) + '_' + str(hour) + '_' + str(minute) + str(second) + '.txt'
            #filenamenb = 'NumberIssues' + str(day) + str(hour) + str(minute) + str(second) + '.txt'
            filenameGI = 'GCListSplit' + str(day) + str(hour) + str(minute) + str(second) + '.xlsx'
          

            #print filename

            #f = open(filenamenb, 'w')



            


          

             
           # print >> f
            #print >> f, "Terminology error (term/string ID):"
            #print >> f, "Terminology error (term/string ID):"
            #print >> f
            
 # Create a xlsxwriter workbook and add a worksheet.
            #workbookxw = xlsxwriter.Workbook(filenamexw)
            #worksheetxw = workbookxw.add_worksheet()
            wbgi = Workbook()

            wsgi1 = wbgi.active
            wsgi1.title = "GCListSplit"

 

            #rowxw = 1

            nu=1


                        
                    

            #row_countgl = sheetgl.get_highest_row()+1
            #row_countst = sheetst.get_highest_row()+1



            

            row_countgl = sheetgl.max_row+1
            row_countst = sheetst.max_row+1   



           


            #for m in range(int(firstline), int(row_countgl)):
            for n in range(int(firstline), int(row_countst)):

                    #eg = sheetgl.cell(row = m,column = ceg).value
                    #tg = sheetgl.cell(row = m,column = ctg).value

                    sid = sheetst.cell(row = n,column = csid).value
                    es = sheetst.cell(row = n,column = ces).value
                    es = str(es)
                    ts = sheetst.cell(row = n,column = cts).value
                    ts = str(ts)

                    #sidl = sheetst.cell(row = n,column = csid).value.lower()
                    #esl = sheetst.cell(row = n,column = ces).value.lower()
                    #tsl = sheetst.cell(row = n,column = cts).value.lower()



                    
                    if es != None:
                        es = unicode(es)
                        esl = es.lower()
                        #tsl = tsl.replace('\n', ' ').replace('\r', '')
                        esls = esl.split()
                    if es == None:
                        esl = "string missing"









                                       
                    #for n in range(int(firstline), int(row_countst)):
                    for m in range(int(firstline), int(row_countgl)):

          
                        #sid = sheetst.cell(row = n,column = csid).value
                        #es = sheetst.cell(row = n,column = ces).value
                        #ts = sheetst.cell(row = n,column = cts).value                  

                        eg = sheetgl.cell(row = m,column = ceg).value
                        eg = str(eg)
                        tg = sheetgl.cell(row = m,column = ctg).value
                        tg = str(tg)

                        #egl = sheetgl.cell(row = m,column = ceg).value.lower()
                        #tgl = sheetgl.cell(row = m,column = ctg).value.lower()


                            
                        if eg != None:
                            eg = unicode(eg)
                            egl = eg.lower()
                            #esl = esl.replace('\n', ' ').replace('\r', '')
                            egls = egl.split()
                        if eg == None:
                            esl = "string missing"

                           
                       
                        
                        

                      
                        #if sheet.cell(m,0).value in sheet.cell(n,2).value and sheet.cell(m,1).value not in sheet.cell(n,3).value:
                        #if eg in es and tg not in ts and ts != "":
                        #if unicode(eg) in unicode(es) and unicode(tg) not in unicode(ts) and unicode(ts) != unicode(""):
                        #if unicode(eg) in unicode(es) and unicode(tg) not in unicode(ts) or unicode(tg) in unicode(ts) and unicode(eg) not in unicode(es):

                        
                        #if unicode(egl) in unicode(esl) and unicode(tgl) not in unicode(tsl) and unicode(tgl)!= "None":

                        #if unicode(eg) in unicode(es) and unicode(tg) not in unicode(ts) and unicode(tg)!= "None":




                        if any(x in egls for x in esls if len(x)>3):

                        #if re.search(r'\b' + unicode(eg) + r'\b', unicode(es)) and unicode(tg) not in unicode(ts) and unicode(tg)!= "None":

                        #if re.search(r'\b' + unicode(eg) + r'\b', unicode(es),  re.UNICODE) and unicode(tg) not in unicode(ts) and unicode(tg)!= "None":



                


                                    nu=nu+1
                      
                        



                                    #worksheetxw.write(nu, 1, eg)
                                    #worksheetxw.write(nu, 2, tg)
                                    #worksheetxw.write(nu, 3, sid)
                                    #worksheetxw.write(nu, 4, es)
                                    #worksheetxw.write(nu, 5, ts)


                                    wsgi1.cell(row=nu, column=1).value = eg
                                    wsgi1.cell(row=nu, column=2).value = tg
                                    wsgi1.cell(row=nu, column=3).value = sid
                                    wsgi1.cell(row=nu, column=4).value = es
                                    wsgi1.cell(row=nu, column=5).value = ts


                                    sys.stdout.write("\rGCListSplit data analysis in progress. This can take several minutes... ")
                                    sys.stdout.flush()



                                    ###print eg



            #worksheetxw.write(1, 1, 'eg')
            #worksheetxw.write(1, 2, 'tg')
            #worksheetxw.write(1, 3, 'sid')
            #worksheetxw.write(1, 4, 'es')
            #worksheetxw.write(1, 5, 'ts')
            
            #worksheetxw.set_default_row(hide_unused_rows=True)
                                
            #workbookxw.close()

            wsgi1.cell(row=1, column=1).value = 'Source Glossary'
            wsgi1.cell(row=1, column=2).value = 'Loc. Glossary'
            wsgi1.cell(row=1, column=3).value = 'String ID'
            wsgi1.cell(row=1, column=4).value = 'Source String'
            wsgi1.cell(row=1, column=5).value = 'Loc. String'

            wbgi.save(filename = filenameGI)

            print "Done"

            #print 'Terminology check done'

            master = Tk()

            w = Label(master, text="\n\n    The output file ('GCListSplit... .xlsx') is in the GlossaryCheck folder.    \n\n", bg="green")
            w.pack()

            mainloop()
            






















          
    def GCList(self):


        
  
     
  

        #ftypes = [('Excel files', '.xls'),('Excel XML', '*.xlsx'), ('All files', '*')]
        ftypes = [('Excel files', '.xlsx')]
        dlg = tkFileDialog.Open(self, filetypes = ftypes)
        fl = dlg.show()



        if fl != '':
            filename = fl
            book = load_workbook(filename)
            sheetgl = book.worksheets[sheetgloss]
            sheetst = book.worksheets[sheetstrng]
         


            year = datetime.datetime.now().year
            month = datetime.datetime.now().month
            hour = datetime.datetime.now().hour
            minute = datetime.datetime.now().minute
            day = datetime.datetime.now().day
            second = datetime.datetime.now().second
    

            #filename = str (year) + str(month)
            #filename = str(year) + '_' + str(month) + '_' + str(day) + '_' + str(hour) + '_' + str(minute) + str(second) + '.txt'
            #filenamenb = 'NumberIssues' + str(day) + str(hour) + str(minute) + str(second) + '.txt'
            filenameGL = 'GCList' + str(day) + str(hour) + str(minute) + str(second) + '.xlsx'
          

            #print filename

            #f = open(filenamenb, 'w')



            


          

             
           # print >> f
            #print >> f, "Terminology error (term/string ID):"
            #print >> f, "Terminology error (term/string ID):"
            #print >> f
            
 # Create a xlsxwriter workbook and add a worksheet.
            #workbookxw = xlsxwriter.Workbook(filenamexw)
            #worksheetxw = workbookxw.add_worksheet()


            wbgl = Workbook()

            wsgl1 = wbgl.active
            wsgl1.title = "GCList"



            #rowxw = 1

            nu=1


                        
                    

            #row_countgl = sheetgl.get_highest_row()+1
            #row_countst = sheetst.get_highest_row()+1




            row_countgl = sheetgl.max_row+1
            row_countst = sheetst.max_row+1              




           


            #for m in range(int(firstline), int(row_countgl)):
            for n in range(int(firstline), int(row_countst)):

                    #eg = sheetgl.cell(row = m,column = ceg).value
                    #tg = sheetgl.cell(row = m,column = ctg).value

                    sid = sheetst.cell(row = n,column = csid).value
                    es = sheetst.cell(row = n,column = ces).value
                    es = str(es)
                    ts = sheetst.cell(row = n,column = cts).value



                    #sidl = sheetst.cell(row = n,column = csid).value.lower()
                    #esl = es.lower()
                    #tsl = sheetst.cell(row = n,column = cts).value.lower()



                    #if es != None and es.isdigit() == False:
                    if es != None: 
                        esl = es.lower()
                        
                        esl = esl.replace('\n', ' ').replace('\r', '')
                    #if es == None or es.isdigit() == True:
                    if es == None:   
                        esl = "string missing"

                    
                   # if ts != None:                        
                       # tsl = ts.lower()
                        #tsl = tsl.replace('\n', ' ').replace('\r', '')
                    #if ts == None:
                       # tsl = "string missing"

                    

                                       
                    #for n in range(int(firstline), int(row_countst)):
                    for m in range(int(firstline), int(row_countgl)):

          
                        #sid = sheetst.cell(row = n,column = csid).value
                        #es = sheetst.cell(row = n,column = ces).value
                        #ts = sheetst.cell(row = n,column = cts).value                  

                        eg = sheetgl.cell(row = m,column = ceg).value
                        eg = str(eg)
                        tg = sheetgl.cell(row = m,column = ctg).value


                        #if eg != None and "." not in eg and eg.isdigit() == False:
                        if eg != None and "." not in eg: 
                            eg = sheetgl.cell(row = m,column = ceg).value.strip()
                            egl = eg.lower()
                        if eg != None and "." in eg:
                            egl = eg
                        #if eg == None or eg.isdigit() == True:
                        if eg == None:
                            eg = "None"
                            egl = "string missing"
                            
                        #if tg != None and "." not in tg:  
                          #  tg = sheetgl.cell(row = m,column = ctg).value.strip()
                           # tgl = tg.lower()
                       # if tg != None and "." in tg:
                           # tgl = eg                         
                       # if tg == None:
                           # tg = "None"
                           # tgl = "string missing"
                            

              
                        





                        #remove leading/trailing spaces

                        #if eg != None:
                            #eg = sheetgl.cell(row = m,column = ceg).value.strip()
                       # if eg == None:
                            #eg = "missing string"
                        #if tg != None:  
                           # tg = sheetgl.cell(row = m,column = ctg).value.strip()
                        #if tg == None:
                            #tg = "missing string"
                            

                        #egl = eg.lower()
                        #tgl = sheetgl.cell(row = m,column = ctg).value

                       
                       
                        
                        

                      
                        #if sheet.cell(m,0).value in sheet.cell(n,2).value and sheet.cell(m,1).value not in sheet.cell(n,3).value:
                        #if eg in es and tg not in ts and ts != "":
                        #if unicode(eg) in unicode(es) and unicode(tg) not in unicode(ts) and unicode(ts) != unicode(""):
                        #if unicode(eg) in unicode(es) and unicode(tg) not in unicode(ts) or unicode(tg) in unicode(ts) and unicode(eg) not in unicode(es):

                        
                        if re.search(r'\b' + unicode(egl) + r'\b', unicode(esl),  re.UNICODE):
                        
                        #if unicode(egl) in unicode(esl):



                                    nu=nu+1
                      
                        



                                    #worksheetxw.write(nu, 1, eg)
                                    #worksheetxw.write(nu, 2, tg)
                                    #worksheetxw.write(nu, 3, sid)
                                    #worksheetxw.write(nu, 4, es)
                                    #worksheetxw.write(nu, 5, ts)



                                    wsgl1.cell(row=nu, column=1).value = eg
                                    wsgl1.cell(row=nu, column=2).value = tg
                                    wsgl1.cell(row=nu, column=3).value = sid
                                    wsgl1.cell(row=nu, column=4).value = es
                                    wsgl1.cell(row=nu, column=5).value = ts

                                    
                                    sys.stdout.write("\rGCList data analysis in progress. This can take several minutes... ")
                                    sys.stdout.flush()



                               

                                    ###print eg



            #worksheetxw.write(1, 1, 'eg')
            #worksheetxw.write(1, 2, 'tg')
            #worksheetxw.write(1, 3, 'sid')
            #worksheetxw.write(1, 4, 'es')
            #worksheetxw.write(1, 5, 'ts')
            
            #worksheetxw.set_default_row(hide_unused_rows=True)
                                
            #workbookxw.close()


            wsgl1.cell(row=1, column=1).value = 'Source Glossary'
            wsgl1.cell(row=1, column=2).value = 'Loc. Glossary'
            wsgl1.cell(row=1, column=3).value = 'String ID'
            wsgl1.cell(row=1, column=4).value = 'Source String'
            wsgl1.cell(row=1, column=5).value = 'Loc. String'

            wbgl.save(filename = filenameGL)


            print "Done"


            

            #print 'Terminology check done'

            master = Tk()

            w = Label(master, text="\n\n    The output file ('GlossaryList... .xlsx') is in the GlossaryCheck folder.    \n\n", bg="green")
            w.pack()

            mainloop()
            

        ####################    
         #####################   
            

          
    def GCListCut(self):


        
  
     
  

        #ftypes = [('Excel files', '.xls'),('Excel XML', '*.xlsx'), ('All files', '*')]
        ftypes = [('Excel files', '.xlsx')]
        dlg = tkFileDialog.Open(self, filetypes = ftypes)
        fl = dlg.show()



        if fl != '':
            filename = fl
            book = load_workbook(filename)
            sheetgl = book.worksheets[sheetgloss]
            sheetst = book.worksheets[sheetstrng]
         


            year = datetime.datetime.now().year
            month = datetime.datetime.now().month
            hour = datetime.datetime.now().hour
            minute = datetime.datetime.now().minute
            day = datetime.datetime.now().day
            second = datetime.datetime.now().second
    

            #filename = str (year) + str(month)
            #filename = str(year) + '_' + str(month) + '_' + str(day) + '_' + str(hour) + '_' + str(minute) + str(second) + '.txt'
            #filenamenb = 'NumberIssues' + str(day) + str(hour) + str(minute) + str(second) + '.txt'
            filenameGL = 'GCListCut' + str(day) + str(hour) + str(minute) + str(second) + '.xlsx'
          

            #print filename

            #f = open(filenamenb, 'w')



            


          

             
           # print >> f
            #print >> f, "Terminology error (term/string ID):"
            #print >> f, "Terminology error (term/string ID):"
            #print >> f
            
 # Create a xlsxwriter workbook and add a worksheet.
            #workbookxw = xlsxwriter.Workbook(filenamexw)
            #worksheetxw = workbookxw.add_worksheet()


            wbgl = Workbook()

            wsgl1 = wbgl.active
            wsgl1.title = "GCListCut"



            #rowxw = 1

            nu=1


                        
                    

            #row_countgl = sheetgl.get_highest_row()+1
            #row_countst = sheetst.get_highest_row()+1




            row_countgl = sheetgl.max_row+1
            row_countst = sheetst.max_row+1              




           


            #for m in range(int(firstline), int(row_countgl)):
            for n in range(int(firstline), int(row_countst)):

                    #eg = sheetgl.cell(row = m,column = ceg).value
                    #tg = sheetgl.cell(row = m,column = ctg).value

                    sid = sheetst.cell(row = n,column = csid).value
                    es = sheetst.cell(row = n,column = ces).value
                    es = str(es)
                    ts = sheetst.cell(row = n,column = cts).value



                    #sidl = sheetst.cell(row = n,column = csid).value.lower()
                    #esl = es.lower()
                    #tsl = sheetst.cell(row = n,column = cts).value.lower()



                    if es != None:                        
                        esl = es.lower()
                        esl = esl.replace('\n', ' ').replace('\r', '')
                    if es == None:
                        esl = "string missing"

                    
                   # if ts != None:                        
                       # tsl = ts.lower()
                        #tsl = tsl.replace('\n', ' ').replace('\r', '')
                    #if ts == None:
                       # tsl = "string missing"

                    

                                       
                    #for n in range(int(firstline), int(row_countst)):
                    for m in range(int(firstline), int(row_countgl)):

          
                        #sid = sheetst.cell(row = n,column = csid).value
                        #es = sheetst.cell(row = n,column = ces).value
                        #ts = sheetst.cell(row = n,column = cts).value                  

                        eg = sheetgl.cell(row = m,column = ceg).value
                        eg = str(eg)
                        tg = sheetgl.cell(row = m,column = ctg).value


                        if eg != None and "." not in eg:
                            eg = sheetgl.cell(row = m,column = ceg).value.strip()
                            egl = eg.lower()
                        if eg != None and "." in eg:
                            egl = eg
                        if eg == None:
                            eg = "None"
                            egl = "string missing"
                            
                        #if tg != None and "." not in tg:  
                          #  tg = sheetgl.cell(row = m,column = ctg).value.strip()
                           # tgl = tg.lower()
                       # if tg != None and "." in tg:
                           # tgl = eg                         
                       # if tg == None:
                           # tg = "None"
                           # tgl = "string missing"
                            

              
                        





                        #remove leading/trailing spaces

                        #if eg != None:
                            #eg = sheetgl.cell(row = m,column = ceg).value.strip()
                       # if eg == None:
                            #eg = "missing string"
                        #if tg != None:  
                           # tg = sheetgl.cell(row = m,column = ctg).value.strip()
                        #if tg == None:
                            #tg = "missing string"
                            

                        #egl = eg.lower()
                        #tgl = sheetgl.cell(row = m,column = ctg).value

                       
                       
                        
                        

                      
                        #if sheet.cell(m,0).value in sheet.cell(n,2).value and sheet.cell(m,1).value not in sheet.cell(n,3).value:
                        #if eg in es and tg not in ts and ts != "":
                        #if unicode(eg) in unicode(es) and unicode(tg) not in unicode(ts) and unicode(ts) != unicode(""):
                        #if unicode(eg) in unicode(es) and unicode(tg) not in unicode(ts) or unicode(tg) in unicode(ts) and unicode(eg) not in unicode(es):

                        
                        #if re.search(r'\b' + unicode(egl) + r'\b', unicode(esl),  re.UNICODE):
                        
                        if unicode(egl) in unicode(esl):



                                    nu=nu+1
                      
                        



                                    #worksheetxw.write(nu, 1, eg)
                                    #worksheetxw.write(nu, 2, tg)
                                    #worksheetxw.write(nu, 3, sid)
                                    #worksheetxw.write(nu, 4, es)
                                    #worksheetxw.write(nu, 5, ts)



                                    wsgl1.cell(row=nu, column=1).value = eg
                                    wsgl1.cell(row=nu, column=2).value = tg
                                    wsgl1.cell(row=nu, column=3).value = sid
                                    wsgl1.cell(row=nu, column=4).value = es
                                    wsgl1.cell(row=nu, column=5).value = ts

                                    
                                    sys.stdout.write("\rGCListCut data analysis in progress. This can take several minutes... ")
                                    sys.stdout.flush()



                               

                                    ###print eg



            #worksheetxw.write(1, 1, 'eg')
            #worksheetxw.write(1, 2, 'tg')
            #worksheetxw.write(1, 3, 'sid')
            #worksheetxw.write(1, 4, 'es')
            #worksheetxw.write(1, 5, 'ts')
            
            #worksheetxw.set_default_row(hide_unused_rows=True)
                                
            #workbookxw.close()


            wsgl1.cell(row=1, column=1).value = 'Source Glossary'
            wsgl1.cell(row=1, column=2).value = 'Loc. Glossary'
            wsgl1.cell(row=1, column=3).value = 'String ID'
            wsgl1.cell(row=1, column=4).value = 'Source String'
            wsgl1.cell(row=1, column=5).value = 'Loc. String'

            wbgl.save(filename = filenameGL)


            print "Done"


            

            #print 'Terminology check done'

            master = Tk()

            w = Label(master, text="\n\n    The output file ('GCListCut... .xlsx') is in the GlossaryCheck folder.    \n\n", bg="green")
            w.pack()

            mainloop()
            

                    
        
#####################
   #####################     

    def GCCreator(self):


        
  
     
  

        #ftypes = [('Excel files', '.xls'),('Excel XML', '*.xlsx'), ('All files', '*')]
        ftypes = [('Excel files', '.xlsx')]
        dlg = tkFileDialog.Open(self, filetypes = ftypes)
        fl = dlg.show()



        if fl != '':
            filename = fl
            book = load_workbook(filename)
            sheetgl = book.worksheets[sheetgloss]
            sheetst = book.worksheets[sheetstrng]
         


            year = datetime.datetime.now().year
            month = datetime.datetime.now().month
            hour = datetime.datetime.now().hour
            minute = datetime.datetime.now().minute
            day = datetime.datetime.now().day
            second = datetime.datetime.now().second
    

            #filename = str (year) + str(month)
            #filename = str(year) + '_' + str(month) + '_' + str(day) + '_' + str(hour) + '_' + str(minute) + str(second) + '.txt'
            #filenamenb = 'NumberIssues' + str(day) + str(hour) + str(minute) + str(second) + '.txt'
            filenameGM = 'GlossaryMat' + str(day) + str(hour) + str(minute) + str(second) + '.xlsx'
          

            #print filename

            #f = open(filenamenb, 'w')



            


          

             
           # print >> f
            #print >> f, "Terminology error (term/string ID):"
            #print >> f, "Terminology error (term/string ID):"
            #print >> f
            
 # Create a xlsxwriter workbook and add a worksheet.
            #workbookxw = xlsxwriter.Workbook(filenamexw)
            #worksheetxw = workbookxw.add_worksheet()


 
            wbgm = Workbook()

            wsgm1 = wbgm.active
            #wsgm1.title = "GlossaryMaterial"

            #rowxw = 1

            


            #row_countgl = sheetgl.get_highest_row()+1
            #row_countst = sheetst.get_highest_row()+1


            
            row_countgl = sheetgl.max_row+1
            row_countst = sheetst.max_row+1   






##############wordfrq_ES


            nuz=1


                        
                    



            

            esl = []

            for zlere in range(int(firstline), int(row_countst)):


                    
                    es = unicode(sheetst.cell(row = zlere,column = ces).value)
                    es = str(es)
                    es = re.sub(r'[^\w\s]','',unicode(es), re.UNICODE)
                    #es = re.sub(r'\b' + unicode(es) + r'\b')


                    
                    esp = es.split()
              
                    esl.insert(0, esp)
            

                    
          
 
          
            mergede = list(itertools.chain(*esl))

                                

          
            countse = Counter(mergede)

# Sort the dictionary by value
            lste = list()
            
            for keye, vale in countse.items():
                lste.append( (vale, keye) )

            lste.sort(reverse=True)

            for keye, vale in lste:
              

                nuz=nuz+1

                #worksheetxw.write(nuz, 1, keye)
                #worksheetxw.write(nuz, 2, vale)

                wsgm1.cell(row=nuz, column=2).value = keye
                wsgm1.cell(row=nuz, column=1).value = vale
            
            

######################wordfrq_TS


                
            #nuzt=1


                        
                    




            

            #tsl = []

            #for zlert in range(int(firstline), int(row_countst)):


                    
                    #ts = unicode(sheetst.cell(row = zlert,column = cts).value)
                                    
                    #ts = str(ts)
                    #ts = re.sub(r'[^\w\s]','',unicode(ts), re.UNICODE)

                    ##ts = re.sub(r'\b' + unicode(ts) + r'\b')
                    
                   # tsp = ts.split()
          
                
                    #tsl.insert(0, tsp)
             

                    
      
          
            #mergedt = list(itertools.chain(*tsl))
                    

          
            #countst = Counter(mergedt)

# Sort the dictionary by value
            #lstt = list()
            
            #for keyt, valt in countst.items():
                #lstt.append( (valt, keyt) )

           # lstt.sort(reverse=True)

            #for keyt, valt in lstt:
                

                #nuzt=nuzt+1

               # #worksheetxw.write(nuzt, 4, keyt)
               # #worksheetxw.write(nuzt, 5, valt)

                #wsgm1.cell(row=nuzt, column=4).value = keyt
                #wsgm1.cell(row=nuzt, column=5).value = valt


                sys.stdout.write("\rWord Freqency data analysis in progress. This can take several minutes... ")
                sys.stdout.flush()
                                    
                      
                        


#######################


            #worksheetxw.write(1, 1, 'keye')
            #worksheetxw.write(1, 2, 'vale')
            #worksheetxw.write(1, 4, 'keyt')
            #worksheetxw.write(1, 5, 'valt')
 
            
            #worksheetxw.set_default_row(hide_unused_rows=True)
                                
            #workbookxw.close()


            wsgm1.cell(row=1, column=2).value = 'Frequency'
            wsgm1.cell(row=1, column=1).value = 'Source term'
            #wsgm1.cell(row=1, column=4).value = 'Frequency'
            #wsgm1.cell(row=1, column=5).value = 'Translation term'


            wbgm.save(filename = filenameGM)


            print "Done"


            

 

            master = Tk()

            w = Label(master, text="\n\n Word Freqency data collected. The output file ('GlossaryMat... .xlsx') is in the GlossaryCheck folder.  \nYou can use this data to create or extend a terminology list.    \n\n", bg="green")
            w.pack()

            mainloop()
            




          
            
            
            

                    

           

            
                
    def NumberCheck(self):
            


  
     
        ftypes = [('Excel files', '.xlsx')]
        dlg = tkFileDialog.Open(self, filetypes = ftypes)
        fl = dlg.show()



        if fl != '':
            filename = fl
          

            book = load_workbook(filename)
            sheetn = book.worksheets[sheetstrng]


            year = datetime.datetime.now().year
            month = datetime.datetime.now().month
            hour = datetime.datetime.now().hour
            minute = datetime.datetime.now().minute
            day = datetime.datetime.now().day
            second = datetime.datetime.now().second
    
            filenameNC = 'NumberCheck' + str(day) + str(hour) + str(minute) + str(second) + '.xlsx'
           



            


          

        
            #workbooknb = xlsxwriter.Workbook(filenamenb)
            #worksheetnb = workbooknb.add_worksheet()

            wbnc = Workbook()

            wsnc1 = wbnc.active
            wsnc1.title = "NumberCheck"            




            nunc=1
            
            #row_countn = sheetn.get_highest_row()+1

            row_countn = sheetn.max_row+1

                       
      
           



            for row_idx in range(int(firstline), int(row_countn)):
                  
                   
                 


                                                          
            
        

                        sid = sheetn.cell(row = row_idx,column = csid).value
                        es = sheetn.cell(row = row_idx,column = ces).value                        
                        #es = str(es)
                        ts = sheetn.cell(row = row_idx,column = cts).value
                        #ts = str(ts)
       
                        #zelle = sheetn.cell(row = row_idx,column = ces).value
                        #zelnex = sheetn.cell(row = row_idx,column = cts).value

                        zelle = unicode(sheetn.cell(row = row_idx,column = ces).value)

                        zelnex = unicode(sheetn.cell(row = row_idx,column = cts).value)
                        



                        #zelnex = re.sub('[^0-9]', '', zelnex)
                        #zelle = re.sub('[^0-9]', '', zelle)
                        #if zelle > 0 and zelle != zelnex and zelnex != '':


                        zelnexNR = re.sub('[^0-9]', '', zelnex)
                        zelleNR = re.sub('[^0-9]', '', zelle)
                        
                        if zelleNR > 0 and zelleNR != zelnexNR and zelnexNR != '':
                              
                            
                                
                               



                                
                                #worksheetnb.write(row_idx-1, 1, sid)
                                #worksheetnb.write(row_idx-1, 2, es)
                                #worksheetnb.write(row_idx-1, 3, ts)

                                
                                nunc=nunc+1


                                wsnc1.cell(row=nunc, column=1).value = sid
                                wsnc1.cell(row=nunc, column=2).value = es
                                wsnc1.cell(row=nunc, column=3).value = ts

                                sys.stdout.write("\rGCNumberCheck data analysis in progress. This can take several minutes... ")
                                sys.stdout.flush()
                         
                               



                                    
            #worksheetnb.set_default_row(hide_unused_rows=True)
                                
            #workbooknb.close()


            wsnc1.cell(row=1, column=1).value = 'String ID'
            wsnc1.cell(row=1, column=2).value = 'Source String'
            wsnc1.cell(row=1, column=3).value = 'Translation'

            wbnc.save(filename = filenameNC)

          

            print 'Done'

              

            master = Tk()

            w = Label(master, text="\n\n    The output file ('NumberCeck... .xlsx') is in the GlossaryCheck folder.    \n\n", bg="green")
            w.pack()

            mainloop()





    def Manual(self):

          

        master = Tk()

        w = Label(master, text="\n\n    Manual created. The Manual file ('Manual.txt') is in the GlossaryCheck folder.     \n\n", bg="green")
        w.pack()
        
        with open("Manual.txt",'w') as gcman:
            
            gcman.write("GlossaryCheck Manual")
            gcman.write("\n\n")
            gcman.write("CONTENT:\n\n")
            gcman.write("INTRODUCTION" + "\n")
            gcman.write("THE INPUT FILE" + "\n")
            gcman.write("GlossaryCheck" + "\n")
            gcman.write("GC MatchCase" + "\n")
            gcman.write("GCList" + "\n")
            gcman.write("GCList (Split)" + "\n")
            gcman.write("GCList (Cut)" + "\n")
            gcman.write("GlossaryList" + "\n")
            gcman.write("NumberCheck" + "\n")
            gcman.write("GlossaryCreator" + "\n")
            
            gcman.write("\n\n")
            gcman.write("INTRODUCTION" + "\n\n")
            gcman.write("GlossaryCheck is a linguistic tool to help find terminology errors in large string based localization projects using spreadsheet files. Terminology translation errors can make texts more difficult to understand or can change the meaning. Without Computer-assisted translation software, it is difficult and time consuming to find terminology translation errors. Correct terminology is a feature of high quality translations, especially of more complex texts (e.g. philosophy or scientific texts, but also novels or texts in games). Imagine a special term in a philosophical text has several different translations..." + "\n\n")
            gcman.write("GlossaryCheck works with bulk inputs of glossaries (up to 1048576 entries) and string files (up to 1048576 strings) via .xlsx spreadsheets. GlossaryCheck is independent of other Computer-assisted translation software and can be used as an analyzing tool, analyzing the bulk output of several translators, collected in one single spreadsheet (e.g. selective output of large MySql databases). The output of GlossaryCheck is a .xlsx spreadsheet with detailed info about localization terminology errors (String ID, terminology term in source and localized language, original string, localized string)." + "\n\n")
            gcman.write("GlossaryCheck lists strings, where the source string contains a terminology term and the translated string does not contain the corresponding term. GlossaryCheck's input tool allows different sensitivity settings e.g. case insensitive/sensitive, word boundaries insensitive/sensitive for thousands of Glossary terms at the same time." + "\n\n")
            gcman.write("GlossaryCheck contains also tools to list all strings containing Terminology terms (GlossaryList), to check typos of numbers inside of strings and to create or extend terminology lists (GCCreator). " + "\n\n")
            gcman.write("GlossaryCheck is written by A.D.Klumpp using Python and the Python library openpyxl including jdcal and et_xmlfile (see license texts below or in the folders of the libraries). GlossaryCheck is released under the terms of the GNU General Public License (See http://www.gnu.org/licenses/). Copyright (C) 2015 A.D.Klumpp. GlossaryCheck is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY. The full copyright notices and the full license texts shall be included in all copies or substantial portions of the Software. Alternative versions of GlossaryCheck are written by A.D. Klumpp using Python and the Python libraries openpyxl (including jdcal and et_xmlfile), XlsxWriter and kivy (thanks to python-for-android). " + "\n\n")
            gcman.write("Python is released under the Python Software Foundation License (see https://www.python.org/download/releases/2.7.6/license/). Openpyxl is released under MIT/Expat license (see https://openpyxl.readthedocs.org/en/latest/). Kivy is released under the MIT License (see https://github.com/kivy/kivy/blob/master/LICENSE). jdcal is released under BSD (see https://pypi.python.org/pypi/jdcal). et_xmlfile is released under MIT (Home-page: https://bitbucket.org/openpyxl/et_xmlfile)." + "\n\n")
            gcman.write("Please read the full license texts Online or in the LICENSES.txt document, which is inside the GlossaryCheck folder." + "\n\n")
            gcman.write("" + "\n\n")

            
            gcman.write("\n\n")
            gcman.write("THE INPUT FILE" + "\n\n")
            gcman.write("Format:.xlsx" + "\n\n")
            gcman.write("Structure:\n\n")

            gcman.write("Sheet 1 (Glossary/Terminology):" + "\n")
            gcman.write("Column 1: Glossary Source language" + "\n")
            gcman.write("Column 2: Glossary Translation" + "\n\n")
            gcman.write("Sheet 2 (Strings):" + "\n")
            gcman.write("Column 1: String ID" + "\n")
            gcman.write("Column 2: Strings Source language" + "\n")
            gcman.write("Column 3: Strings Translation" + "\n\n")
            gcman.write("Don't input formatting data" + "\n")
            gcman.write("For some analysis programs not all input columns are required. See below." + "\n")
            gcman.write("Open the input file (open the START-Menu and select the analysis program, see below) and wait until the done-pop up appears. This can take several minutes." + "\n")
            gcman.write("Open with LibreOffice, OpenOffice or MS Excel the output file, which is now in the same location as the GlossaryCheck starter file." + "\n")
            gcman.write("" + "\n")
            gcman.write("" + "\n")
             
        
            gcman.write("\n\n")
            gcman.write("GlossaryCheck" + "\n")
            gcman.write("Input: All input columns required." + "\n")
            gcman.write("Lists strings, where the source string contains a glossary term and the translated string does not contain the corresponding glossary term. Case insensitive, but sensitive to word boundaries (if the term is 'Def' it will not search for 'Definition', but for 'def'). Ignores line breaks inside of Excel cells. The analysis of large files can take longer, depending on the performance of the system)." + "\n")

            gcman.write("\n\n")
            gcman.write("GC MatchCase" + "\n")
            gcman.write("Input: All input columns required." + "\n")
            gcman.write("All input columns required. Same features as GlossaryCheck, but case sensitive." + "\n")

        
            gcman.write("\n\n")
            gcman.write("GCList" + "\n")
            gcman.write("Input: Not required: Glossary Translation." + "\n")
            gcman.write("Lists all strings containing Glossary terms. Case insensitive, word boundaries sensitive, ignores line breaks in strings." + "\n")           

            gcman.write("\n\n")
            gcman.write("GCList (Split)" + "\n")
            gcman.write("Input: Not required: Glossary Translation." + "\n")
            gcman.write("Splits Glossary terms, which consists of several longer words so that the parts are new Glossary terms." + "\n")           

            gcman.write("\n\n")
            gcman.write("GCList (Cut)" + "\n")
            gcman.write("Input: All input columns required." + "\n")
            gcman.write("Lists Strings, where the beginning of a word is the same as a Glossary term (if the term is 'Def' it will also search for 'Definition'). Case and word boundaries insensitive, ignores line breaks in strings." + "\n")

            gcman.write("\n\n")
            gcman.write("NumberCheck" + "\n")
            gcman.write("Input: Tab 1 can be let empty. Please always put the strings on Tab 2." + "\n")
            gcman.write("Lists all strings, where numeral values (not number words) are different compared to the source string. E.g.: Source string: 130 tanks, translation: 13 Tanks" + "\n")

            gcman.write("\n\n")
            gcman.write("GCCreator" + "\n")
            gcman.write("Input: Tab 1 can be let empty. Please always put the strings on Tab 2." + "\n")
            gcman.write("Returns a word frequency count list (Column 1: term, Column 2: frequency), sorted by frequency. This data can be used in order to create a glossary list or to extend an existing glossary list, by selecting the most important words. In order to list the used translations, copy the selection into Column 1, tab 1 of the GCList input file (let the second Column of tab 1 empty) and start GCList. With the GCList output the second column of tab 1 (glossary translation) of the GC-Input file can be completed in order to finish the GC-input file. Please note that for the word frequency counter only English is supported as source language." + "\n")
                
            gcman.write("\n\n")
            gcman.write("\n\n")
        
            #string = StringProperty('')

            


        mainloop()









    def Legal(self):


        master = Tk()

        w = Label(master, text="\n\n    LICENSES text created. The LICENSES file ('LICENSES.txt') is in the GlossaryCheck folder.    \n\n", bg="green")
        w.pack()
        
        with open("LICENSES.txt",'w') as gpl:
            gpl.write("CONTENT:\n")
            gpl.write(GCLH + "\n")
            gpl.write(OPLH + "\n")
            gpl.write(JDCALH + "\n")
            gpl.write(ETXMLLH + "\n")
            #gpl.write(KVLH + "\n")
            gpl.write(GPLH + "\n")
            gpl.write(PLH + "\n")
            
            gpl.write("\n\n")
            gpl.write(GCLH + "\n")
            gpl.write(GCLT + "\n")
            
            gpl.write("\n\n")
            gpl.write(OPLH + "\n")
            gpl.write(OPLT + "\n")
            
            gpl.write("\n\n")
            gpl.write(JDCALH + "\n")
            gpl.write(JDCALT + "\n")

            gpl.write("\n\n")
            gpl.write(ETXMLLH + "\n")
            gpl.write(ETXMLLT + "\n")

            #gpl.write("\n\n")
            #gpl.write(KVLH + "\n")
            #gpl.write(KVLT + "\n")

            gpl.write("\n\n")
            gpl.write(GPLH + "\n")
            gpl.write(GPLT + "\n")           

            gpl.write("\n\n")
            gpl.write(PLH + "\n")
            gpl.write(PLT + "\n")
                
            gpl.write("\n\n")
        
            #string = StringProperty('')

            


        mainloop()




            
                    

def main():

  


    root = Tk()
    ex = GCMain(root)
    root.geometry("700x180+500+500")
    root.mainloop()  


if __name__ == '__main__':
    main()
