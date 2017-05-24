#!/usr/bin/env python

# -*- coding: utf-8 -*-


#Search+Mark

#Requirements:
#openpyxl 2.3.3.
#jdcal 1.2
#et_xmlfile-1.0.1
#Python 2.7 (2.7.10)

#Licenses


GCLH = "Search+Mark LICENSE"
GCLT= "Search+Mark is written by A.D. Klumpp using Python and the Python library openpyxl including jdcal and et_xmlfile (see license texts below or in the folders of the libraries). Search+Mark is released under the terms of the GNU General Public License. Copyright (C) 2016 A.D.Klumpp. Search+Mark is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY. The full copyright notices and the full license texts shall be included in all copies or substantial portions of the Software."

OPLH = "OPENPYXL 2.3.3 LICENSE"
OPLT = "(http://openpyxl.readthedocs.org/en/latest/_modules/openpyxl/worksheet/header_footer.html) Copyright (c) 2010-2015 openpyxl. Permission is hereby granted, free of charge, to any person obtaining a copy of this software and associated documentation files (the 'Software'), to deal in the Software without restriction, including without limitation the rights to use, copy, modify, merge, publish, distribute, sublicense, and/or sell copies of the Software, and to permit persons to whom the Software is furnished to do so, subject to the following conditions: The above copyright notice and this permission notice shall be included in all copies or substantial portions of the Software. THE SOFTWARE IS PROVIDED 'AS IS', WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE."

JDCALH = "jdcal 1.2 LICENSE"
JDCALT = '(https://pypi.python.org/pypi/jdcal) Copyright (c) 2011, Prasanth Nair. All rights reserved.\nRedistribution and use in source and binary forms, with or without modification, are permitted provided that the following conditions are met:\n1. Redistributions of source code must retain the above copyright notice, this list of conditions and the following disclaimer.\n2. Redistributions in binary form must reproduce the above copyright notice, this list of conditions and the following disclaimer in the documentation and/or other materials provided with the distribution.\nTHIS SOFTWARE IS PROVIDED BY THE COPYRIGHT HOLDERS AND CONTRIBUTORS "AS IS" AND ANY EXPRESS OR IMPLIED WARRANTIES, INCLUDING, BUT NOT LIMITED TO, THE IMPLIED WARRANTIES OF MERCHANTABILITY AND FITNESS FOR A PARTICULAR PURPOSE ARE DISCLAIMED. IN NO EVENT SHALL THE COPYRIGHT HOLDER OR CONTRIBUTORS BE LIABLE FOR ANY DIRECT, INDIRECT, INCIDENTAL, SPECIAL, EXEMPLARY, OR CONSEQUENTIAL DAMAGES (INCLUDING, BUT NOT LIMITED TO, PROCUREMENT OF SUBSTITUTE GOODS OR SERVICES; LOSS OF USE, DATA, OR PROFITS; OR BUSINESS INTERRUPTION) HOWEVER CAUSED AND ON ANY THEORY OF LIABILITY, WHETHER IN CONTRACT, STRICT LIABILITY, OR TORT (INCLUDING NEGLIGENCE OR OTHERWISE) ARISING IN ANY WAY OUT OF THE USE OF THIS SOFTWARE, EVEN IF ADVISED OF THE POSSIBILITY OF SUCH DAMAGE.'

ETXMLLH = "et_xmlfile"
ETXMLLT = "et_xmlfile is a low memory library for creating large XML files. It is based upon the xmlfile module from lxml <http://lxml.de/api.html#incremental-xml-generation>_ with the aim of allowing code to be developed that will work with both libraries. It was developed initially for the openpyxl project but is now a standalone module. The code was written by Elias Rabel as part of the Python Duesseldorf <http://pyddf.de>_ openpyxl sprint in September 2014. Version: 1.0.1. License: MIT. Home-page: https://bitbucket.org/openpyxl/et_xmlfile"

#KVLH = "KIVY LICENSE"
#KVLT = "(https://github.com/kivy/kivy/blob/master/LICENSE) Copyright (c) 2010-2015 Kivy Team and other contributors. Permission is hereby granted, free of charge, to any person obtaining a copy of this software and associated documentation files (the 'Software'), to deal in the Software without restriction, including without limitation the rights to use, copy, modify, merge, publish, distribute, sublicense, and/or sell copies of the Software, and to permit persons to whom the Software is furnished to do so, subject to the following conditions: The above copyright notice and this permission notice shall be included in all copies or substantial portions of the Software. THE SOFTWARE IS PROVIDED 'AS IS', WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE."

GPLH = "GNU GENERAL PUBLIC LICENSE"
GPLT = "See http://www.gnu.org/licenses/"

PLH = "PYTHON LICENSE"
PLT = "See https://www.python.org/download/releases/2.7.6/license/"




#sheet of input .xlsx
# 0=1, 1=2...

sheetgloss = 0
#sheet glossary

sheetstrng = 1
#sheet strings


firstline = 1

firstcol = 1




#import xlrd
import re
#from xlrd import open_workbook,cellname, XL_CELL_TEXT, cellnameabs, colname
from Tkinter import *
import tkFileDialog
import Tkinter, tkSimpleDialog
from tkFileDialog import askopenfilename
from Tkinter import Frame, Tk, BOTH, Text, Menu, END
import subprocess as sub
import tkMessageBox
import Tkinter
import pickle
import datetime
#import xlsxwriter
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.cell import get_column_letter

from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell



from collections import Counter
import itertools
import sys
reload(sys)
sys.setdefaultencoding('utf-8')




class GCMain(Frame):




   

    def __init__(self, parent):
        Frame.__init__(self, parent)   

       
        self.parent = parent        
        self.initUI()

        frame = Frame(parent)
        frame.pack()

        


        Label(parent, text="Search+Mark\n\n\n").pack()
        
   
        Label(parent, text="For further info please read the Manual and the LICENSE-texts (START-Menu).\n\n\n\n").pack()


            


    def initUI(self):

        self.parent.title('Search+Mark')
        self.pack(fill=BOTH, expand=1)

        menubar = Menu(self.parent)
        self.parent.config(menu=menubar)

                    

        
        fileMenu = Menu(menubar)
      
       
        menubar.add_cascade(label="START", menu=fileMenu)

       

        fileMenu.add_command(label="Open Path File for:", command=None)
        fileMenu.add_command(label="Search+Mark: Specific (Exact)", command=self.SMOne)
        fileMenu.add_command(label="Search+Mark: Specific (Fuzzy)", command=self.SMOneF)
        fileMenu.add_command(label="Search+Mark: Universal (Exact)", command=self.SMAll)
        fileMenu.add_command(label="Search+Mark: Universal (Fuzzy)", command=self.SMAllF)
      
        fileMenu.add_command(label="Search+Edit: Universal (Exact)", command=self.SEAll)
        fileMenu.add_command(label="Search+Edit: Universal (Fuzzy)", command=self.SEAllF)
        fileMenu.add_command(label="----------------------------------", command=None)
        fileMenu.add_command(label="Manual", command=self.Manual)  
        fileMenu.add_command(label="Legal/Licenses", command=self.Legal)




    def SMAll(self):


        #while True:

        


            #ki = raw_input('Search for: ')

            
            #print ki

            

           
#########open path excel file######



        ftypes = [('Excel files', '.xlsx')]
        dlg = tkFileDialog.Open(self, filetypes = ftypes)
        fl = dlg.show()

     



        if fl != '':


                
            filename = fl
                #filename = ptif
            bookp = load_workbook(filename)
            sheetp = bookp.worksheets[0]
                #sheetst = book.worksheets[sheetstrng]

#########open path excel file End######  



############Excel EOD file#################


            year = datetime.datetime.now().year
            month = datetime.datetime.now().month
            hour = datetime.datetime.now().hour
            minute = datetime.datetime.now().minute
            day = datetime.datetime.now().day
            second = datetime.datetime.now().second
        

                #filename = str (year) + str(month)
                #filename = str(year) + '_' + str(month) + '_' + str(day) + '_' + str(hour) + '_' + str(minute) + str(second) + '.txt'
                #filenamenb = 'NumberIssues' + str(day) + str(hour) + str(minute) + str(second) + '.txt'
            filenameGI = 'Report' + str(day) + str(hour) + str(minute) + str(second) + '.xlsx'
              



                #f = open(filenamenb, 'w')


                
     # Create a  workbook and add a worksheet.
                #workbookxw = xlsxwriter.Workbook(filenamexw)
                #worksheetxw = workbookxw.add_worksheet()
            wbgi = Workbook()

            wsgi1 = wbgi.active
            wsgi1.title = "Found"

     

                #rowxw = 1

            nu=1


############Excel EOD file End#################                           


############Constant inputs##################

            #keycol = 1

            #markcol = 7

            #comcol = 8

            #marktext = "R"


            #colFill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')



            #keycol = raw_input('Enter Number of Search-Column, e.g. "1" for first Column: ')
            #keycol = int(keycol)
            #print keycol

            print "Search+Mark"
            print
            print "Mode:"
            print "Universal (Search in all columns)," 
            print "Exact match (word boundaries sensitive, case insenesitive)."
            print "Don't start another mode, while program is running."
            print "Restart the program in order to change between modes."
            print

            markcol = raw_input('Enter Number of Mark-Column, e.g. for first column "1": ')            
            markcol = int(markcol)
            #print markcol
                
            marktext = raw_input('Enter Mark-Text, e.g. "Checked": ')
            marktext = unicode(marktext)
            #print marktext

            mcolor = raw_input('Change the color of the Mark-Cell, y/n? ')
            if mcolor == "y":
                color = raw_input('Enter the color of the cell, "r" for red, "y" for yellow, "b" for blue, "g" for green: ')
                if color == "r":
                    colFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                if color == "y":
                    colFill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')                    
                if color == "b":
                    colFill = PatternFill(start_color='FF0000FF', end_color='FF0000FF', fill_type='solid')
                if color == "g":
                        colFill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
            if mcolor == "n":
               #colFill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
               colFill = "None"

                     
                    

            comques = raw_input('Do you use a column with comments, y/n? ')

            if comques == "y":
                comcol = raw_input('Enter Number of Comment-Column, eg. "1": ')
                comcol = int(comcol)
            
                    
            if comques == "n":

                comcol = "None"
                    
                print "Comments disabled."

                         






    #DARKRED = 'FF800000'
    #DARKBLUE = 'FF000080'
    #DARKGREEN = 'FF008000'
    #DARKYELLOW = 'FF808000' 



 
                       
###########Search input######################


                
            while True:

                ki = raw_input('Search for: ')

                kil = ki.lower()

            
                #print ki


###########Search input End######################


#############Path file##################

                nf = 1

                row_countp = sheetp.max_row+1

                for pr in range(int(firstline), int(row_countp)):
                    
                    ptf = unicode(sheetp.cell(row = pr,column = 1).value)

                    ptf = unicode(ptf)

                    #print ptf

                    if ptf != None:
                        ptf = sheetp.cell(row = pr,column = 1).value.strip()
                    if ptf == None:
                        ptf = "No"

                    #print ptf
                        

                    if ptf != None and ptf != "No":

                        bookkey = load_workbook(ptf)



#######################################


            

#########MAX sheet for input key files###########

                #for sheet in book.worksheets:
                    #print sheet

                    for idx,sheet in enumerate(bookkey.worksheets):
                        idx=idx
                    #print idx

                    maxsheet = idx
                    maxsheet = maxsheet + 1
                #print "maxsheet"
                #print maxsheet

#########max sheet for input files end##########






                #row_countgl = sheetgl.get_highest_row()+1
                #row_countst = sheetst.get_highest_row()+1


               
                #row_countgl = sheetgl.max_row+1
                #row_countst = sheetst.max_row+1

                #col_countgl = sheetgl.max_column+1

                    

                    for sn in range(int(0), int(maxsheet)):
                    
                        

                        #print "sn"
                        #print sn
                        sheetv = bookkey.worksheets[sn]

                        #print sheetg

                        row_countv = sheetv.max_row+1
                       
                        col_countv = sheetv.max_column+1



                        
                         
                        for co in range(int(firstline), int(col_countv)):

                            co=co

                            #print "co"
                            #print co


                            for m in range(int(firstline), int(row_countv)):
                                    
                                #for n in range(int(firstline), int(row_countst)):

                                        
                    
                                        #textpr = "In progress..."
                                        #sys.stdout.write(str(textpr))
                                        #sys.stdout.flush()

                                  

                                    k = sheetv.cell(row = m,column = co).value


                                    if k != None:
                                        k = str(k)
                                        k = str(sheetv.cell(row = m,column = co).value).strip()
                                        k = k.lower()
                                        k = k.replace('\n', ' ').replace('\r', '')
                                    if k == None:
                                        k = "None"

                                    
                                        #n1 = sheetgl.cell(row = m,column = 2).value
                                        #r1 = sheetgl.cell(row = m,column = 3).value
                                   

                                    #if sn == 0 and co ==1:
                                        #print "k sn0"
                                        #print k
                                        #print ki
                                        

                                    #if unicode(kil) in unicode(k):

                                    if unicode(kil)== unicode(k):
                                       # print k, m, co, sheetgl
                                        #print "True"
                                        print ("Found '%s' in file %s,"%(k, ptf))
                                        print ("%s, col. %s, row %s"%(sheetv, co, m))
                                        print ("Content of row %s: "%(m))

                                        for cellrow in range(int(firstcol), int(col_countv)):
                                            printcell = sheetv.cell(row = m,column = cellrow).value
                                            printcell = unicode(printcell)
                                            print ("Col. '%s': %s "%(cellrow, printcell))
                                        
                                        #print k, m, co, sheetv
                                        #if unicode(ki) in unicode(n1):
                                          #  print n1, m, sheetgl
                                       # if unicode(ki) in unicode(r1):
                                          #  print r1, m, sheetgl

                                        nf = nf + 1


                                        qi = raw_input('Update file y/n? ')



                                        if qi=="y":



                                            #redFill = PatternFill(start_color='FFFF0000',
                                               #end_color='FFFF0000',
                                               #fill_type='solid')

                                            if colFill != "None":

                                                sheetv.cell(row=m, column=markcol).fill = colFill

                                            mf = sheetv.cell(row=m, column=markcol).value

                                            if mf != None:
                                                print "Content in mark cell found:"
                                                print mf
                                                ea = raw_input("Edit anyway, y/n? ")
                                                if ea=="y":
                                                    sheetv.cell(row=m, column=markcol).value = marktext
                                                    print "Mark cell updated."
                                                    upd = "Yes"
                                                if ea=="n":
                                                    mf = sheetv.cell(row=m, column=markcol).value
                                                    print "Mark cell not updated."
                                                    upd = "No"

                                            if mf == None:
                                                sheetv.cell(row=m, column=markcol).value = marktext
                                                print "Mark cell updated."
                                                upd = "Yes"
                                                

                                            

                                            if comcol != "None":
                                                cf = sheetv.cell(row=m, column=comcol).value
                                                if cf != None:
                                                    print "Content in comment cell found:"
                                                    print cf
                                                    eac = raw_input("Edit anyway, y/n? ")
                                                    if eac=="y":
                                                        comtext = raw_input('Enter comment: ')
                                                        comtext = unicode(comtext)
                                                        sheetv.cell(row=m, column=comcol).value = comtext
                                                        print "Comment cell updated."
                                                        upd = "Yes"
                                                    if eac=="n":
                                                        #cf = sheetv.cell(row=m, column=comcol).value
                                                        print "Comment cell not updated."                                                   
                                                        upd = "No"

                                                if cf == None:
                                                    comtext = raw_input('Enter comment: ')
                                                    comtext = unicode(comtext)
                                                    sheetv.cell(row=m, column=comcol).value = comtext
                                                    print "Comment cell updated."
                                                    upd = "Yes"

                                                
                                                #comtext = raw_input('Enter comment: ')
                                                #comtext = unicode(comtext)
                                                #sheetv.cell(row=m, column=comcol).value = comtext
                                            
                                            bookkey.save(filename = ptf)
                                            #print "File updated"
                                            #upd = "Yes"

                                        if qi=="n":
                                            upd = "No"
                                            print "File not updated."                    
  


                                        nu=nu+1

                                        wsgi1.cell(row=1, column=1).value = 'Input'
                                        wsgi1.cell(row=1, column=2).value = 'Found item'
                                        wsgi1.cell(row=1, column=3).value = 'File'
                                        wsgi1.cell(row=1, column=4).value = 'Sheet'
                                        wsgi1.cell(row=1, column=5).value = 'Col.'
                                        wsgi1.cell(row=1, column=6).value = 'Row'
                                        wsgi1.cell(row=1, column=7).value = 'File updated?'
                                        
                                        

                                        wsgi1.cell(row=nu, column=1).value = ki
                                        wsgi1.cell(row=nu, column=2).value = k
                                        wsgi1.cell(row=nu, column=3).value = ptf
                                        wsgi1.cell(row=nu, column=4).value = unicode(sheetv)
                                        wsgi1.cell(row=nu, column=5).value = co
                                        wsgi1.cell(row=nu, column=6).value = m
                                        wsgi1.cell(row=nu, column=7).value = upd
                                        

                                        wbgi.save(filename = filenameGI)

                if nf < 2:
                    print "Not found."
                                     


###################sm all Fuzzy###########

    def SMAllF(self):


        #while True:

        


            #ki = raw_input('Search for: ')

            
            #print ki

            

           
#########open path excel file######



        ftypes = [('Excel files', '.xlsx')]
        dlg = tkFileDialog.Open(self, filetypes = ftypes)
        fl = dlg.show()

     



        if fl != '':


                
            filename = fl
                #filename = ptif
            bookp = load_workbook(filename)
            sheetp = bookp.worksheets[0]
                #sheetst = book.worksheets[sheetstrng]

#########open path excel file End######  



############Excel EOD file#################


            year = datetime.datetime.now().year
            month = datetime.datetime.now().month
            hour = datetime.datetime.now().hour
            minute = datetime.datetime.now().minute
            day = datetime.datetime.now().day
            second = datetime.datetime.now().second
        

                #filename = str (year) + str(month)
                #filename = str(year) + '_' + str(month) + '_' + str(day) + '_' + str(hour) + '_' + str(minute) + str(second) + '.txt'
                #filenamenb = 'NumberIssues' + str(day) + str(hour) + str(minute) + str(second) + '.txt'
            filenameGI = 'Report' + str(day) + str(hour) + str(minute) + str(second) + '.xlsx'
              



                #f = open(filenamenb, 'w')


                
     # Create a  workbook and add a worksheet.
                #workbookxw = xlsxwriter.Workbook(filenamexw)
                #worksheetxw = workbookxw.add_worksheet()
            wbgi = Workbook()

            wsgi1 = wbgi.active
            wsgi1.title = "Found"

     

                #rowxw = 1

            nu=1


############Excel EOD file End#################                           


############Constant inputs##################

            #keycol = 1

            #markcol = 7

            #comcol = 8

            #marktext = "R"


            #colFill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')



            #keycol = raw_input('Enter Number of Search-Column, e.g. "1" for first Column: ')
            #keycol = int(keycol)
            #print keycol

            print "Search+Mark"
            print
            print "Mode:"
            print "Universal (Search in all columns),"
            print "Fuzzy (word boundaries insensitive, case insenesitive)."               
            print "Don't open another mode while program is running."
            print "Restart the program in order to change between modes."
            print

            markcol = raw_input('Enter Number of Mark-Column, e.g. for first column "1": ')            
            markcol = int(markcol)
            #print markcol
                
            marktext = raw_input('Enter Mark-Text, e.g. "Checked": ')
            marktext = unicode(marktext)
            #print marktext

            mcolor = raw_input('Change the color of the Mark-Cell, y/n? ')
            if mcolor == "y":
                color = raw_input('Enter the color of the cell, "r" for red, "y" for yellow, "b" for blue, "g" for green: ')
                if color == "r":
                    colFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                if color == "y":
                    colFill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')                    
                if color == "b":
                    colFill = PatternFill(start_color='FF0000FF', end_color='FF0000FF', fill_type='solid')
                if color == "g":
                        colFill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
            if mcolor == "n":
               #colFill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
               colFill = "None"

                     
                    

            comques = raw_input('Do you use a column with comments, y/n? ')

            if comques == "y":
                comcol = raw_input('Enter Number of Comment-Column, eg. "1": ')
                comcol = int(comcol)
            
                    
            if comques == "n":

                comcol = "None"
                    
                print "Comments disabled."

                         






    #DARKRED = 'FF800000'
    #DARKBLUE = 'FF000080'
    #DARKGREEN = 'FF008000'
    #DARKYELLOW = 'FF808000' 



 
                       
###########Search input######################


                
            while True:

                ki = raw_input('Search for: ')

                kil = ki.lower()

            
                #print ki


###########Search input End######################


#############Path file##################

                nf = 1

                row_countp = sheetp.max_row+1

                for pr in range(int(firstline), int(row_countp)):
                    
                    ptf = unicode(sheetp.cell(row = pr,column = 1).value)

                    ptf = unicode(ptf)

                    #print ptf

                    if ptf != None:
                        ptf = sheetp.cell(row = pr,column = 1).value.strip()
                    if ptf == None:
                        ptf = "No"

                    #print ptf
                        

                    if ptf != None and ptf != "No":

                        bookkey = load_workbook(ptf)



#######################################


            

#########MAX sheet for input key files###########

                #for sheet in book.worksheets:
                    #print sheet

                    for idx,sheet in enumerate(bookkey.worksheets):
                        idx=idx
                    #print idx

                    maxsheet = idx
                    maxsheet = maxsheet + 1
                #print "maxsheet"
                #print maxsheet

#########max sheet for input files end##########






                #row_countgl = sheetgl.get_highest_row()+1
                #row_countst = sheetst.get_highest_row()+1


               
                #row_countgl = sheetgl.max_row+1
                #row_countst = sheetst.max_row+1

                #col_countgl = sheetgl.max_column+1

                    

                    for sn in range(int(0), int(maxsheet)):
                    
                        

                        #print "sn"
                        #print sn
                        sheetv = bookkey.worksheets[sn]

                        #print sheetg

                        row_countv = sheetv.max_row+1
                       
                        col_countv = sheetv.max_column+1



                        
                         
                        for co in range(int(firstline), int(col_countv)):

                            co=co

                            #print "co"
                            #print co


                            for m in range(int(firstline), int(row_countv)):
                                    
                                #for n in range(int(firstline), int(row_countst)):

                                        
                    
                                        #textpr = "In progress..."
                                        #sys.stdout.write(str(textpr))
                                        #sys.stdout.flush()

                                  

                                    k = sheetv.cell(row = m,column = co).value


                                    if k != None:
                                        k = str(k)
                                        k = str(sheetv.cell(row = m,column = co).value).strip()
                                        k = k.lower()
                                        k = k.replace('\n', ' ').replace('\r', '')
                                    if k == None:
                                        k = "None"

                                    
                                        #n1 = sheetgl.cell(row = m,column = 2).value
                                        #r1 = sheetgl.cell(row = m,column = 3).value
                                   

                                    #if sn == 0 and co ==1:
                                        #print "k sn0"
                                        #print k
                                        #print ki
                                        

                                    if unicode(kil) in unicode(k):

                                    #if unicode(kil)== unicode(k):
                                       # print k, m, co, sheetgl
                                        #print "True"
                                        print ("Found '%s' in file %s,"%(k, ptf))
                                        print ("%s, col. %s, row %s"%(sheetv, co, m))
                                        print ("Content of row %s: "%(m))

                                        for cellrow in range(int(firstcol), int(col_countv)):
                                            printcell = sheetv.cell(row = m,column = cellrow).value
                                            printcell = unicode(printcell)
                                            print ("Col. '%s': %s "%(cellrow, printcell))
                                        
                                        #print k, m, co, sheetv
                                        #if unicode(ki) in unicode(n1):
                                          #  print n1, m, sheetgl
                                       # if unicode(ki) in unicode(r1):
                                          #  print r1, m, sheetgl

                                        nf = nf + 1


                                        qi = raw_input('Update file y/n? ')



                                        if qi=="y":



                                            #redFill = PatternFill(start_color='FFFF0000',
                                               #end_color='FFFF0000',
                                               #fill_type='solid')

                                            if colFill != "None":

                                                sheetv.cell(row=m, column=markcol).fill = colFill

                                            mf = sheetv.cell(row=m, column=markcol).value

                                            if mf != None:
                                                print "Content in mark cell found:"
                                                print mf
                                                ea = raw_input("Edit anyway, y/n? ")
                                                if ea=="y":
                                                    sheetv.cell(row=m, column=markcol).value = marktext
                                                    print "Mark cell updated."
                                                    upd = "Yes"
                                                if ea=="n":
                                                    mf = sheetv.cell(row=m, column=markcol).value
                                                    print "Mark cell not updated."
                                                    upd = "No"

                                            if mf == None:
                                                sheetv.cell(row=m, column=markcol).value = marktext
                                                print "Mark cell updated."
                                                upd = "Yes"
                                                

                                            

                                            if comcol != "None":
                                                cf = sheetv.cell(row=m, column=comcol).value
                                                if cf != None:
                                                    print "Content in comment cell found:"
                                                    print cf
                                                    eac = raw_input("Edit anyway, y/n? ")
                                                    if eac=="y":
                                                        comtext = raw_input('Enter comment: ')
                                                        comtext = unicode(comtext)
                                                        sheetv.cell(row=m, column=comcol).value = comtext
                                                        print "Comment cell updated."
                                                        upd = "Yes"
                                                    if eac=="n":
                                                        #cf = sheetv.cell(row=m, column=comcol).value
                                                        print "Comment cell not updated."                                                   
                                                        upd = "No"

                                                if cf == None:
                                                    comtext = raw_input('Enter comment: ')
                                                    comtext = unicode(comtext)
                                                    sheetv.cell(row=m, column=comcol).value = comtext
                                                    print "Comment cell updated."
                                                    upd = "Yes"

                                                
                                                #comtext = raw_input('Enter comment: ')
                                                #comtext = unicode(comtext)
                                                #sheetv.cell(row=m, column=comcol).value = comtext
                                            
                                            bookkey.save(filename = ptf)
                                            #print "File updated"
                                            #upd = "Yes"

                                        if qi=="n":
                                            upd = "No"
                                            print "File not updated."                    
  


                                        nu=nu+1

                                        wsgi1.cell(row=1, column=1).value = 'Input'
                                        wsgi1.cell(row=1, column=2).value = 'Found item'
                                        wsgi1.cell(row=1, column=3).value = 'File'
                                        wsgi1.cell(row=1, column=4).value = 'Sheet'
                                        wsgi1.cell(row=1, column=5).value = 'Col.'
                                        wsgi1.cell(row=1, column=6).value = 'Row'
                                        wsgi1.cell(row=1, column=7).value = 'File updated?'
                                        
                                        

                                        wsgi1.cell(row=nu, column=1).value = ki
                                        wsgi1.cell(row=nu, column=2).value = k
                                        wsgi1.cell(row=nu, column=3).value = ptf
                                        wsgi1.cell(row=nu, column=4).value = unicode(sheetv)
                                        wsgi1.cell(row=nu, column=5).value = co
                                        wsgi1.cell(row=nu, column=6).value = m
                                        wsgi1.cell(row=nu, column=7).value = upd
                                        

                                        wbgi.save(filename = filenameGI)

                if nf < 2:
                    print "Not found."
                                     




 

            
################sm all fuzzy end############
                                       



######################Search + Mark################
###################################################




    def SMOne(self):


        #while True:

        


            #ki = raw_input('Search for: ')

            
            #print ki

            

           
#########open path excel file######       
      

        ftypes = [('Excel files', '.xlsx')]
        dlg = tkFileDialog.Open(self, filetypes = ftypes)
        fl = dlg.show()

            




        if fl != '':

                
            filename = fl
                #filename = ptif
            bookp = load_workbook(filename)
            sheetp = bookp.worksheets[0]
                #sheetst = book.worksheets[sheetstrng]

#########open path excel file End######  



############Excel EOD file#################


            year = datetime.datetime.now().year
            month = datetime.datetime.now().month
            hour = datetime.datetime.now().hour
            minute = datetime.datetime.now().minute
            day = datetime.datetime.now().day
            second = datetime.datetime.now().second
        

                #filename = str (year) + str(month)
                #filename = str(year) + '_' + str(month) + '_' + str(day) + '_' + str(hour) + '_' + str(minute) + str(second) + '.txt'
                #filenamenb = 'NumberIssues' + str(day) + str(hour) + str(minute) + str(second) + '.txt'
            filenameGI = 'Report' + str(day) + str(hour) + str(minute) + str(second) + '.xlsx'
              



                #f = open(filenamenb, 'w')


                
     # Create a  workbook and add a worksheet.
                #workbookxw = xlsxwriter.Workbook(filenamexw)
                #worksheetxw = workbookxw.add_worksheet()
            wbgi = Workbook()

            wsgi1 = wbgi.active
            wsgi1.title = "Found"

     

                #rowxw = 1

            nu=1


############Excel EOD file End#################                           

############Constant inputs##################

            #keycol = 1

            #markcol = 7

            #comcol = 8

            #marktext = "R"


            #colFill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')

            print "Search+Mark"
            print 
            print "Mode:"
            print "Specific (Search in one specific column),"
            print "Exact (word boundaries sensitive, case insenesitive)."
            print "Don't open another mode while program is running."
            print "Restart the program in order to change between modes."
            print



            keycol = raw_input('Enter Number of Search-Column, e.g. "1" for first Column: ')
            keycol = int(keycol)
            #print keycol
     

            markcol = raw_input('Enter Number of Mark-Column, e.g. "5": ')            
            markcol = int(markcol)
            #print markcol
                
            marktext = raw_input('Enter Mark-Text, e.g. "Checked": ')
            marktext = unicode(marktext)
            #print marktext

            mcolor = raw_input('Change the color of the Mark-Cell, y/n? ')
            if mcolor == "y":
                color = raw_input('Enter the color of the cell, "r" for red, "y" for yellow, "b" for blue, "g" for green: ')
                if color == "r":
                    colFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                if color == "y":
                    colFill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')                    
                if color == "b":
                    colFill = PatternFill(start_color='FF0000FF', end_color='FF0000FF', fill_type='solid')
                if color == "g":
                        colFill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
            if mcolor == "n":
               #colFill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
               colFill = "None"

                     
                    

            comques = raw_input('Do you use a column with comments, y/n? ')

            if comques == "y":
                comcol = raw_input('Enter Number of Comment-Column, eg. "1": ')
                comcol = int(comcol)
            
                    
            if comques == "n":

                comcol = "None"
                    
                print "Comments disabled."

                         






    #DARKRED = 'FF800000'
    #DARKBLUE = 'FF000080'
    #DARKGREEN = 'FF008000'
    #DARKYELLOW = 'FF808000' 
                       
###########Search input######################


                
            while True:

                ki = raw_input('Search for: ')
                kil = ki.lower()

            
                #print ki


###########Search input End######################


#############Path file##################

                nf = 1

                row_countp = sheetp.max_row+1

                for pr in range(int(firstline), int(row_countp)):
                    
                    ptf = unicode(sheetp.cell(row = pr,column = 1).value)

                    ptf = unicode(ptf)

                    #print ptf

                    if ptf != None:
                        ptf = sheetp.cell(row = pr,column = 1).value.strip()
                    if ptf == None:
                        ptf = "No"

                    #print ptf
                        

                    if ptf != None and ptf != "No":

                        bookkey = load_workbook(ptf)



#######################################


            

#########MAX sheet for input key files###########

                #for sheet in book.worksheets:
                    #print sheet

                    for idx,sheet in enumerate(bookkey.worksheets):
                        idx=idx
                    #print idx

                    maxsheet = idx
                    maxsheet = maxsheet + 1
                #print "maxsheet"
                #print maxsheet

#########max sheet for input files end##########






                #row_countgl = sheetgl.get_highest_row()+1
                #row_countst = sheetst.get_highest_row()+1


               
                #row_countgl = sheetgl.max_row+1
                #row_countst = sheetst.max_row+1

                #col_countgl = sheetgl.max_column+1

                    #nf = 1

                    for sn in range(int(0), int(maxsheet)):

                        
                        
                    
                        

                        #print "sn"
                        #print sn
                        sheetv = bookkey.worksheets[sn]

                        #print sheetg

                        row_countv = sheetv.max_row+1
                       
                        col_countv = sheetv.max_column+1



                        
                         
                        #for co in range(int(firstline), int(col_countv)):

                            #co=co

                            #print "co"
                            #print co


                        for m in range(int(firstline), int(row_countv)):
                                    
                                #for n in range(int(firstline), int(row_countst)):

                                        
                    
                                        #textpr = "In progress..."
                                        #sys.stdout.write(str(textpr))
                                        #sys.stdout.flush()

                                    #keycol = 1

                                    k = sheetv.cell(row = m,column = keycol).value


                                    if k != None:
                                        k = unicode(k)
                                        k = str(sheetv.cell(row = m,column = keycol).value).strip()
                                        k = k.lower()
                                        
     
                                    if k == None:
                                        k = "None"

                                        
                                        #n1 = sheetgl.cell(row = m,column = 2).value
                                        #r1 = sheetgl.cell(row = m,column = 3).value

                                    #if sn == 0 and co ==1:
                                        #print "k sn0"
                                        #print k
                                        #print ki
                                        

                                    #if unicode(ki) in unicode(k):
                                    #ki = k.lower()
                                    if unicode(kil)== unicode(k):
                                       # print k, m, co, sheetgl
                                        #print "True"
                                        print ("Found '%s' in file %s,"%(k, ptf))
                                        print ("%s, col. %s, row %s"%(sheetv, keycol, m))
                                        print ("Content of row %s: "%(m))

                                        for cellrow in range(int(firstcol), int(col_countv)):
                                            printcell = sheetv.cell(row = m,column = cellrow).value
                                            printcell = unicode(printcell)
                                            print ("Col. '%s': %s "%(cellrow, printcell))
                                        
                                        #print k, m, co, sheetv
                                        #if unicode(ki) in unicode(n1):
                                          #  print n1, m, sheetgl
                                       # if unicode(ki) in unicode(r1):
                                          #  print r1, m, sheetgl

                                        nf = nf + 1


                                        qi = raw_input('Update file y/n? ')



                                        if qi=="y":



                                            #redFill = PatternFill(start_color='FFFF0000',
                                               #end_color='FFFF0000',
                                               #fill_type='solid')

                                            if colFill != "None":

                                                sheetv.cell(row=m, column=markcol).fill = colFill

                                            mf = sheetv.cell(row=m, column=markcol).value

                                            if mf != None:
                                                print "Content in mark cell found:"
                                                print mf
                                                ea = raw_input("Edit anyway, y/n? ")
                                                if ea=="y":
                                                    sheetv.cell(row=m, column=markcol).value = marktext
                                                    print "Mark cell updated."
                                                    upd = "Yes"
                                                if ea=="n":
                                                    mf = sheetv.cell(row=m, column=markcol).value
                                                    print "Mark cell not updated."
                                                    upd = "No"

                                            if mf == None:
                                                sheetv.cell(row=m, column=markcol).value = marktext
                                                print "Mark cell updated."
                                                upd = "Yes"
                                                

                                            

                                            if comcol != "None":
                                                cf = sheetv.cell(row=m, column=comcol).value
                                                if cf != None:
                                                    print "Content in comment cell found:"
                                                    print cf
                                                    eac = raw_input("Edit anyway, y/n? ")
                                                    if eac=="y":
                                                        comtext = raw_input('Enter comment: ')
                                                        comtext = unicode(comtext)
                                                        sheetv.cell(row=m, column=comcol).value = comtext
                                                        print "Comment cell updated."
                                                        upd = "Yes"
                                                    if eac=="n":
                                                        #cf = sheetv.cell(row=m, column=comcol).value
                                                        print "Comment cell not updated."                                                   
                                                        upd = "No"

                                                if cf == None:
                                                    comtext = raw_input('Enter comment: ')
                                                    comtext = unicode(comtext)
                                                    sheetv.cell(row=m, column=comcol).value = comtext
                                                    print "Comment cell updated."
                                                    upd = "Yes"

                                                
                                                #comtext = raw_input('Enter comment: ')
                                                #comtext = unicode(comtext)
                                                #sheetv.cell(row=m, column=comcol).value = comtext
                                            
                                            bookkey.save(filename = ptf)
                                            #print "File updated"
                                            #upd = "Yes"

                                        if qi=="n":
                                            upd = "No"
                                            print "File not updated."                    
  


                                        nu=nu+1

                                        wsgi1.cell(row=1, column=1).value = 'Input'
                                        wsgi1.cell(row=1, column=2).value = 'Found item'
                                        wsgi1.cell(row=1, column=3).value = 'File'
                                        wsgi1.cell(row=1, column=4).value = 'Sheet'
                                        wsgi1.cell(row=1, column=5).value = 'Col.'
                                        wsgi1.cell(row=1, column=6).value = 'Row'
                                        wsgi1.cell(row=1, column=7).value = 'File updated?'
                                        
                                        

                                        wsgi1.cell(row=nu, column=1).value = ki
                                        wsgi1.cell(row=nu, column=2).value = k
                                        wsgi1.cell(row=nu, column=3).value = ptf
                                        wsgi1.cell(row=nu, column=4).value = unicode(sheetv)
                                        wsgi1.cell(row=nu, column=5).value = keycol
                                        wsgi1.cell(row=nu, column=6).value = m
                                        wsgi1.cell(row=nu, column=7).value = upd
                                        

                                        wbgi.save(filename = filenameGI)

                if nf < 2:
                    print "Not found."
                                     



                                    
        


                                           


#####################Search + Mark End#############
###################################################




##############smone Fuzzy#########################




    def SMOneF(self):


        #while True:

        


            #ki = raw_input('Search for: ')

            
            #print ki

            

           
#########open path excel file######       
      

        ftypes = [('Excel files', '.xlsx')]
        dlg = tkFileDialog.Open(self, filetypes = ftypes)
        fl = dlg.show()

            




        if fl != '':

                
            filename = fl
                #filename = ptif
            bookp = load_workbook(filename)
            sheetp = bookp.worksheets[0]
                #sheetst = book.worksheets[sheetstrng]

#########open path excel file End######  



############Excel EOD file#################


            year = datetime.datetime.now().year
            month = datetime.datetime.now().month
            hour = datetime.datetime.now().hour
            minute = datetime.datetime.now().minute
            day = datetime.datetime.now().day
            second = datetime.datetime.now().second
        

                #filename = str (year) + str(month)
                #filename = str(year) + '_' + str(month) + '_' + str(day) + '_' + str(hour) + '_' + str(minute) + str(second) + '.txt'
                #filenamenb = 'NumberIssues' + str(day) + str(hour) + str(minute) + str(second) + '.txt'
            filenameGI = 'Report' + str(day) + str(hour) + str(minute) + str(second) + '.xlsx'
              



                #f = open(filenamenb, 'w')


                
     # Create a  workbook and add a worksheet.
                #workbookxw = xlsxwriter.Workbook(filenamexw)
                #worksheetxw = workbookxw.add_worksheet()
            wbgi = Workbook()

            wsgi1 = wbgi.active
            wsgi1.title = "Found"

     

                #rowxw = 1

            nu=1


############Excel EOD file End#################                           

############Constant inputs##################

            #keycol = 1

            #markcol = 7

            #comcol = 8

            #marktext = "R"


            #colFill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')

            print "Search+Mark"
            print
            print "Mode:"
            print "Specific (Search in one specific column),"
            print "Fuzzy (word boundaries insensitive, case insenesitive)."
            print "Don't open another mode while program is running."
            print "Restart the program in order to change between modes."
            print

            keycol = raw_input('Enter Number of Search-Column, e.g. "1" for first Column: ')
            keycol = int(keycol)
            #print keycol
     

            markcol = raw_input('Enter Number of Mark-Column, e.g. "5": ')            
            markcol = int(markcol)
            #print markcol
                
            marktext = raw_input('Enter Mark-Text, e.g. "Checked": ')
            marktext = unicode(marktext)
            #print marktext

            mcolor = raw_input('Change the color of the Mark-Cell, y/n? ')
            if mcolor == "y":
                color = raw_input('Enter the color of the cell, "r" for red, "y" for yellow, "b" for blue, "g" for green: ')
                if color == "r":
                    colFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                if color == "y":
                    colFill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')                    
                if color == "b":
                    colFill = PatternFill(start_color='FF0000FF', end_color='FF0000FF', fill_type='solid')
                if color == "g":
                        colFill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
            if mcolor == "n":
               #colFill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
               colFill = "None"

                     
                    

            comques = raw_input('Do you use a column with comments, y/n? ')

            if comques == "y":
                comcol = raw_input('Enter Number of Comment-Column, eg. "1": ')
                comcol = int(comcol)
            
                    
            if comques == "n":

                comcol = "None"
                    
                print "Comments disabled."

                         






    #DARKRED = 'FF800000'
    #DARKBLUE = 'FF000080'
    #DARKGREEN = 'FF008000'
    #DARKYELLOW = 'FF808000' 
                       
###########Search input######################


                
            while True:

                ki = raw_input('Search for: ')
                kil = ki.lower()

            
                #print ki


###########Search input End######################


#############Path file##################

                nf = 1

                row_countp = sheetp.max_row+1

                for pr in range(int(firstline), int(row_countp)):
                    
                    ptf = unicode(sheetp.cell(row = pr,column = 1).value)

                    ptf = unicode(ptf)

                    #print ptf

                    if ptf != None:
                        ptf = sheetp.cell(row = pr,column = 1).value.strip()
                    if ptf == None:
                        ptf = "No"

                    #print ptf
                        

                    if ptf != None and ptf != "No":

                        bookkey = load_workbook(ptf)



#######################################


            

#########MAX sheet for input key files###########

                #for sheet in book.worksheets:
                    #print sheet

                    for idx,sheet in enumerate(bookkey.worksheets):
                        idx=idx
                    #print idx

                    maxsheet = idx
                    maxsheet = maxsheet + 1
                #print "maxsheet"
                #print maxsheet

#########max sheet for input files end##########






                #row_countgl = sheetgl.get_highest_row()+1
                #row_countst = sheetst.get_highest_row()+1


               
                #row_countgl = sheetgl.max_row+1
                #row_countst = sheetst.max_row+1

                #col_countgl = sheetgl.max_column+1

                    #nf = 1

                    for sn in range(int(0), int(maxsheet)):

                        
                        
                    
                        

                        #print "sn"
                        #print sn
                        sheetv = bookkey.worksheets[sn]

                        #print sheetg

                        row_countv = sheetv.max_row+1
                       
                        col_countv = sheetv.max_column+1



                        
                         
                        #for co in range(int(firstline), int(col_countv)):

                            #co=co

                            #print "co"
                            #print co


                        for m in range(int(firstline), int(row_countv)):
                                    
                                #for n in range(int(firstline), int(row_countst)):

                                        
                    
                                        #textpr = "In progress..."
                                        #sys.stdout.write(str(textpr))
                                        #sys.stdout.flush()

                                    #keycol = 1

                                    k = sheetv.cell(row = m,column = keycol).value


                                    if k != None:
                                        k = unicode(k)
                                        k = str(sheetv.cell(row = m,column = keycol).value).strip()
                                        k = k.lower()
                                        
     
                                    if k == None:
                                        k = "None"

                                        
                                        #n1 = sheetgl.cell(row = m,column = 2).value
                                        #r1 = sheetgl.cell(row = m,column = 3).value

                                    #if sn == 0 and co ==1:
                                        #print "k sn0"
                                        #print k
                                        #print ki
                                        

                                    #if unicode(ki) in unicode(k):
                                    #ki = k.lower()
                                    if unicode(kil) in unicode(k):
                                       # print k, m, co, sheetgl
                                        #print "True"
                                        print ("Found '%s' in file %s,"%(k, ptf))
                                        print ("%s, col. %s, row %s"%(sheetv, keycol, m))
                                        print ("Content of row %s: "%(m))

                                        for cellrow in range(int(firstcol), int(col_countv)):
                                            printcell = sheetv.cell(row = m,column = cellrow).value
                                            printcell = unicode(printcell)
                                            print ("Col. '%s': %s "%(cellrow, printcell))
                                        
                                        #print k, m, co, sheetv
                                        #if unicode(ki) in unicode(n1):
                                          #  print n1, m, sheetgl
                                       # if unicode(ki) in unicode(r1):
                                          #  print r1, m, sheetgl

                                        nf = nf + 1


                                        qi = raw_input('Update file y/n? ')



                                        if qi=="y":



                                            #redFill = PatternFill(start_color='FFFF0000',
                                               #end_color='FFFF0000',
                                               #fill_type='solid')

                                            if colFill != "None":

                                                sheetv.cell(row=m, column=markcol).fill = colFill

                                            mf = sheetv.cell(row=m, column=markcol).value

                                            if mf != None:
                                                print "Content in mark cell found:"
                                                print mf
                                                ea = raw_input("Edit anyway, y/n? ")
                                                if ea=="y":
                                                    sheetv.cell(row=m, column=markcol).value = marktext
                                                    print "Mark cell updated."
                                                    upd = "Yes"
                                                if ea=="n":
                                                    mf = sheetv.cell(row=m, column=markcol).value
                                                    print "Mark cell not updated."
                                                    upd = "No"

                                            if mf == None:
                                                sheetv.cell(row=m, column=markcol).value = marktext
                                                print "Mark cell updated."
                                                upd = "Yes"
                                                

                                            

                                            if comcol != "None":
                                                cf = sheetv.cell(row=m, column=comcol).value
                                                if cf != None:
                                                    print "Content in comment cell found:"
                                                    print cf
                                                    eac = raw_input("Edit anyway, y/n? ")
                                                    if eac=="y":
                                                        comtext = raw_input('Enter comment: ')
                                                        comtext = unicode(comtext)
                                                        sheetv.cell(row=m, column=comcol).value = comtext
                                                        print "Comment cell updated."
                                                        upd = "Yes"
                                                    if eac=="n":
                                                        #cf = sheetv.cell(row=m, column=comcol).value
                                                        print "Comment cell not updated."                                                   
                                                        upd = "No"

                                                if cf == None:
                                                    comtext = raw_input('Enter comment: ')
                                                    comtext = unicode(comtext)
                                                    sheetv.cell(row=m, column=comcol).value = comtext
                                                    print "Comment cell updated."
                                                    upd = "Yes"

                                                
                                                #comtext = raw_input('Enter comment: ')
                                                #comtext = unicode(comtext)
                                                #sheetv.cell(row=m, column=comcol).value = comtext
                                            
                                            bookkey.save(filename = ptf)
                                            #print "File updated"
                                            #upd = "Yes"

                                        if qi=="n":
                                            upd = "No"
                                            print "File not updated."                    
  


                                        nu=nu+1

                                        wsgi1.cell(row=1, column=1).value = 'Input'
                                        wsgi1.cell(row=1, column=2).value = 'Found item'
                                        wsgi1.cell(row=1, column=3).value = 'File'
                                        wsgi1.cell(row=1, column=4).value = 'Sheet'
                                        wsgi1.cell(row=1, column=5).value = 'Col.'
                                        wsgi1.cell(row=1, column=6).value = 'Row'
                                        wsgi1.cell(row=1, column=7).value = 'File updated?'
                                        
                                        

                                        wsgi1.cell(row=nu, column=1).value = ki
                                        wsgi1.cell(row=nu, column=2).value = k
                                        wsgi1.cell(row=nu, column=3).value = ptf
                                        wsgi1.cell(row=nu, column=4).value = unicode(sheetv)
                                        wsgi1.cell(row=nu, column=5).value = keycol
                                        wsgi1.cell(row=nu, column=6).value = m
                                        wsgi1.cell(row=nu, column=7).value = upd
                                        

                                        wbgi.save(filename = filenameGI)

                if nf < 2:
                    print "Not found."
                                     



                                    
        


 

##############smone fuzzy#########################




#################S+edit univ exact###################



    def SEAll(self):


        #while True:

        


            #ki = raw_input('Search for: ')

            
            #print ki

            

           
#########open path excel file######



        ftypes = [('Excel files', '.xlsx')]
        dlg = tkFileDialog.Open(self, filetypes = ftypes)
        fl = dlg.show()

     



        if fl != '':


                
            filename = fl
                #filename = ptif
            bookp = load_workbook(filename)
            sheetp = bookp.worksheets[0]
                #sheetst = book.worksheets[sheetstrng]

#########open path excel file End######  



############Excel EOD file#################


            year = datetime.datetime.now().year
            month = datetime.datetime.now().month
            hour = datetime.datetime.now().hour
            minute = datetime.datetime.now().minute
            day = datetime.datetime.now().day
            second = datetime.datetime.now().second
        

                #filename = str (year) + str(month)
                #filename = str(year) + '_' + str(month) + '_' + str(day) + '_' + str(hour) + '_' + str(minute) + str(second) + '.txt'
                #filenamenb = 'NumberIssues' + str(day) + str(hour) + str(minute) + str(second) + '.txt'
            filenameGI = 'Report' + str(day) + str(hour) + str(minute) + str(second) + '.xlsx'
              



                #f = open(filenamenb, 'w')


                
     # Create a  workbook and add a worksheet.
                #workbookxw = xlsxwriter.Workbook(filenamexw)
                #worksheetxw = workbookxw.add_worksheet()
            wbgi = Workbook()

            wsgi1 = wbgi.active
            wsgi1.title = "Found"

     

                #rowxw = 1

            nu=1


############Excel EOD file End#################                           


############Constant inputs##################

            #keycol = 1

            #markcol = 7

            #comcol = 8

            #marktext = "R"


            #colFill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')



            #keycol = raw_input('Enter Number of Search-Column, e.g. "1" for first Column: ')
            #keycol = int(keycol)
            #print keycol

            print "Search+Edit"
            print
            print "Mode:"
            print "Universal (Search in all columns)," 
            print "Exact match (word boundaries sensitive, case insenesitive)."
            print "Don't start another mode, while program is running."
            print "Restart the program in order to change between modes."
            print

            #markcol = raw_input('Enter Number of Mark-Column, e.g. for first column "1": ')            
            #markcol = int(markcol)
            
                
            #marktext = raw_input('Enter Mark-Text, e.g. "Checked": ')
            #marktext = unicode(marktext)
            

            mcolor = raw_input('Change the color of the Edit-Cell, y/n? ')
            if mcolor == "y":
                color = raw_input('Enter the color of the cell, "r" for red, "y" for yellow, "b" for blue, "g" for green: ')
                if color == "r":
                    colFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                if color == "y":
                    colFill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')                    
                if color == "b":
                    colFill = PatternFill(start_color='FF0000FF', end_color='FF0000FF', fill_type='solid')
                if color == "g":
                        colFill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
            if mcolor == "n":
             
               colFill = "None"

                     
                    

            comques = raw_input('Do you use a column with comments, y/n? ')

            if comques == "y":
                comcol = raw_input('Enter Number of Comment-Column, eg. "1": ')
                comcol = int(comcol)
            
                    
            if comques == "n":

                comcol = "None"
                    
                print "Comments disabled."

                         






    #DARKRED = 'FF800000'
    #DARKBLUE = 'FF000080'
    #DARKGREEN = 'FF008000'
    #DARKYELLOW = 'FF808000' 



 
                       
###########Search input######################


                
            while True:

                ki = raw_input('Search for: ')

                kil = ki.lower()

            
                #print ki


###########Search input End######################


#############Path file##################

                nf = 1

                row_countp = sheetp.max_row+1

                for pr in range(int(firstline), int(row_countp)):
                    
                    ptf = unicode(sheetp.cell(row = pr,column = 1).value)

                    ptf = unicode(ptf)

                    #print ptf

                    if ptf != None:
                        ptf = sheetp.cell(row = pr,column = 1).value.strip()
                    if ptf == None:
                        ptf = "No"

                    #print ptf
                        

                    if ptf != None and ptf != "No":

                        bookkey = load_workbook(ptf)



#######################################


            

#########MAX sheet for input key files###########

                #for sheet in book.worksheets:
                    #print sheet

                    for idx,sheet in enumerate(bookkey.worksheets):
                        idx=idx
                    #print idx

                    maxsheet = idx
                    maxsheet = maxsheet + 1
                #print "maxsheet"
                #print maxsheet

#########max sheet for input files end##########






                #row_countgl = sheetgl.get_highest_row()+1
                #row_countst = sheetst.get_highest_row()+1


               
                #row_countgl = sheetgl.max_row+1
                #row_countst = sheetst.max_row+1

                #col_countgl = sheetgl.max_column+1

                    

                    for sn in range(int(0), int(maxsheet)):
                    
                        

                        #print "sn"
                        #print sn
                        sheetv = bookkey.worksheets[sn]

                        #print sheetg

                        row_countv = sheetv.max_row+1
                       
                        col_countv = sheetv.max_column+1



                        
                         
                        for co in range(int(firstline), int(col_countv)):

                            co=co

                            #print "co"
                            #print co


                            for m in range(int(firstline), int(row_countv)):
                                    
                                #for n in range(int(firstline), int(row_countst)):

                                        
                    
                                        #textpr = "In progress..."
                                        #sys.stdout.write(str(textpr))
                                        #sys.stdout.flush()

                                  

                                    k = sheetv.cell(row = m,column = co).value


                                    if k != None:
                                        k = str(k)
                                        k = str(sheetv.cell(row = m,column = co).value).strip()
                                        k = k.lower()
                                        k = k.replace('\n', ' ').replace('\r', '')
                                    if k == None:
                                        k = "None"

                                    
                                        #n1 = sheetgl.cell(row = m,column = 2).value
                                        #r1 = sheetgl.cell(row = m,column = 3).value
                                   

                                    #if sn == 0 and co ==1:
                                        #print "k sn0"
                                        #print k
                                        #print ki
                                        

                                    #if unicode(kil) in unicode(k):

                                    if unicode(kil)== unicode(k):
                                       # print k, m, co, sheetgl
                                        #print "True"
                                        print ("Found '%s' in file %s,"%(k, ptf))
                                        print ("%s, col. %s, row %s"%(sheetv, co, m))
                                        print ("Content of row %s: "%(m))

                                        for cellrow in range(int(firstcol), int(col_countv)):
                                            printcell = sheetv.cell(row = m,column = cellrow).value
                                            printcell = unicode(printcell)
                                            print ("Col. '%s': %s "%(cellrow, printcell))
                                        
                                        #print k, m, co, sheetv
                                        #if unicode(ki) in unicode(n1):
                                          #  print n1, m, sheetgl
                                       # if unicode(ki) in unicode(r1):
                                          #  print r1, m, sheetgl

                                        nf = nf + 1


                                        qi = raw_input('Edit row y/n? ')



                                        if qi=="y":

                                            editcol = raw_input('Enter Number of Edit-Column, e.g. for first column "1": ')            
                                            editcol = int(editcol)

                                            edtext = raw_input('Enter text for edit: ')
                                            edtext = unicode(edtext)
                                            



                                          

                                            if colFill != "None":

                                                sheetv.cell(row=m, column=editcol).fill = colFill

                                            mf = sheetv.cell(row=m, column=editcol).value

                                            if mf != None:
                                                print "Content in cell found:"
                                                print mf
                                                ea = raw_input("Edit anyway, y/n? ")
                                                if ea=="y":
                                                    sheetv.cell(row=m, column=editcol).value = edtext
                                                    print "Cell updated."
                                                    upd = "Yes"
                                                if ea=="n":
                                                    mf = sheetv.cell(row=m, column=markcol).value
                                                    print "Cell not updated."
                                                    upd = "No"

                                            if mf == None:
                                                sheetv.cell(row=m, column=editcol).value = edtext
                                                print "Cell updated."
                                                upd = "Yes"
                                                

                                            

                                            if comcol != "None":
                                                cf = sheetv.cell(row=m, column=comcol).value
                                                if cf != None:
                                                    print "Content in comment cell found:"
                                                    print cf
                                                    eac = raw_input("Edit anyway, y/n? ")
                                                    if eac=="y":
                                                        comtext = raw_input('Enter comment: ')
                                                        comtext = unicode(comtext)
                                                        sheetv.cell(row=m, column=comcol).value = comtext
                                                        print "Comment cell updated."
                                                        upd = "Yes"
                                                    if eac=="n":
                                                        
                                                        print "Comment cell not updated."                                                   
                                                        upd = "No"

                                                if cf == None:
                                                    comtext = raw_input('Enter comment: ')
                                                    comtext = unicode(comtext)
                                                    sheetv.cell(row=m, column=comcol).value = comtext
                                                    print "Comment cell updated."
                                                    upd = "Yes"

                                                
                                              
                                            
                                            bookkey.save(filename = ptf)
                                            

                                        if qi=="n":
                                            upd = "No"
                                            print "File not updated."                    
  


                                        nu=nu+1

                                        wsgi1.cell(row=1, column=1).value = 'Input'
                                        wsgi1.cell(row=1, column=2).value = 'Found item'
                                        wsgi1.cell(row=1, column=3).value = 'File'
                                        wsgi1.cell(row=1, column=4).value = 'Sheet'
                                        wsgi1.cell(row=1, column=5).value = 'Col.'
                                        wsgi1.cell(row=1, column=6).value = 'Row'
                                        wsgi1.cell(row=1, column=7).value = 'File updated?'
                                        
                                        

                                        wsgi1.cell(row=nu, column=1).value = ki
                                        wsgi1.cell(row=nu, column=2).value = k
                                        wsgi1.cell(row=nu, column=3).value = ptf
                                        wsgi1.cell(row=nu, column=4).value = unicode(sheetv)
                                        wsgi1.cell(row=nu, column=5).value = co
                                        wsgi1.cell(row=nu, column=6).value = m
                                        wsgi1.cell(row=nu, column=7).value = upd
                                        

                                        wbgi.save(filename = filenameGI)

                if nf < 2:
                    print "Not found."
                                     



################se univ exact end########################
            

################se univ fuzzy#########################



    def SEAllF(self):


        #while True:

        


            #ki = raw_input('Search for: ')

            
            #print ki

            

           
#########open path excel file######



        ftypes = [('Excel files', '.xlsx')]
        dlg = tkFileDialog.Open(self, filetypes = ftypes)
        fl = dlg.show()

     



        if fl != '':


                
            filename = fl
                #filename = ptif
            bookp = load_workbook(filename)
            sheetp = bookp.worksheets[0]
                #sheetst = book.worksheets[sheetstrng]

#########open path excel file End######  



############Excel EOD file#################


            year = datetime.datetime.now().year
            month = datetime.datetime.now().month
            hour = datetime.datetime.now().hour
            minute = datetime.datetime.now().minute
            day = datetime.datetime.now().day
            second = datetime.datetime.now().second
        

                #filename = str (year) + str(month)
                #filename = str(year) + '_' + str(month) + '_' + str(day) + '_' + str(hour) + '_' + str(minute) + str(second) + '.txt'
                #filenamenb = 'NumberIssues' + str(day) + str(hour) + str(minute) + str(second) + '.txt'
            filenameGI = 'Report' + str(day) + str(hour) + str(minute) + str(second) + '.xlsx'
              



                #f = open(filenamenb, 'w')


                
     # Create a  workbook and add a worksheet.
                #workbookxw = xlsxwriter.Workbook(filenamexw)
                #worksheetxw = workbookxw.add_worksheet()
            wbgi = Workbook()

            wsgi1 = wbgi.active
            wsgi1.title = "Found"

     

                #rowxw = 1

            nu=1


############Excel EOD file End#################                           


############Constant inputs##################

            #keycol = 1

            #markcol = 7

            #comcol = 8

            #marktext = "R"


            #colFill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')



            #keycol = raw_input('Enter Number of Search-Column, e.g. "1" for first Column: ')
            #keycol = int(keycol)
            #print keycol

            print "Search+Edit"
            print
            print "Mode:"
            print "Universal (Search in all columns)," 
            print "Fuzzy match (word boundaries insensitive, case insenesitive)."
            print "Don't start another mode, while program is running."
            print "Restart the program in order to change between modes."
            print

            #markcol = raw_input('Enter Number of Mark-Column, e.g. for first column "1": ')            
            #markcol = int(markcol)
            
                
            #marktext = raw_input('Enter Mark-Text, e.g. "Checked": ')
            #marktext = unicode(marktext)
            

            mcolor = raw_input('Change the color of the Edit-Cell, y/n? ')
            if mcolor == "y":
                color = raw_input('Enter the color of the cell, "r" for red, "y" for yellow, "b" for blue, "g" for green: ')
                if color == "r":
                    colFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                if color == "y":
                    colFill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')                    
                if color == "b":
                    colFill = PatternFill(start_color='FF0000FF', end_color='FF0000FF', fill_type='solid')
                if color == "g":
                        colFill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
            if mcolor == "n":
             
               colFill = "None"

                     
                    

            comques = raw_input('Do you use a column with comments, y/n? ')

            if comques == "y":
                comcol = raw_input('Enter Number of Comment-Column, eg. "1": ')
                comcol = int(comcol)
            
                    
            if comques == "n":

                comcol = "None"
                    
                print "Comments disabled."

                         






    #DARKRED = 'FF800000'
    #DARKBLUE = 'FF000080'
    #DARKGREEN = 'FF008000'
    #DARKYELLOW = 'FF808000' 



 
                       
###########Search input######################


                
            while True:

                ki = raw_input('Search for: ')

                kil = ki.lower()

            
                #print ki


###########Search input End######################


#############Path file##################

                nf = 1

                row_countp = sheetp.max_row+1

                for pr in range(int(firstline), int(row_countp)):
                    
                    ptf = unicode(sheetp.cell(row = pr,column = 1).value)

                    ptf = unicode(ptf)

                    #print ptf

                    if ptf != None:
                        ptf = sheetp.cell(row = pr,column = 1).value.strip()
                    if ptf == None:
                        ptf = "No"

                    #print ptf
                        

                    if ptf != None and ptf != "No":

                        bookkey = load_workbook(ptf)



#######################################


            

#########MAX sheet for input key files###########

                #for sheet in book.worksheets:
                    #print sheet

                    for idx,sheet in enumerate(bookkey.worksheets):
                        idx=idx
                    #print idx

                    maxsheet = idx
                    maxsheet = maxsheet + 1
                #print "maxsheet"
                #print maxsheet

#########max sheet for input files end##########






                #row_countgl = sheetgl.get_highest_row()+1
                #row_countst = sheetst.get_highest_row()+1


               
                #row_countgl = sheetgl.max_row+1
                #row_countst = sheetst.max_row+1

                #col_countgl = sheetgl.max_column+1

                    

                    for sn in range(int(0), int(maxsheet)):
                    
                        

                        #print "sn"
                        #print sn
                        sheetv = bookkey.worksheets[sn]

                        #print sheetg

                        row_countv = sheetv.max_row+1
                       
                        col_countv = sheetv.max_column+1



                        
                         
                        for co in range(int(firstline), int(col_countv)):

                            co=co

                            #print "co"
                            #print co


                            for m in range(int(firstline), int(row_countv)):
                                    
                                #for n in range(int(firstline), int(row_countst)):

                                        
                    
                                        #textpr = "In progress..."
                                        #sys.stdout.write(str(textpr))
                                        #sys.stdout.flush()

                                  

                                    k = sheetv.cell(row = m,column = co).value


                                    if k != None:
                                        k = str(k)
                                        k = str(sheetv.cell(row = m,column = co).value).strip()
                                        k = k.lower()
                                        k = k.replace('\n', ' ').replace('\r', '')
                                    if k == None:
                                        k = "None"

                                    
                                        #n1 = sheetgl.cell(row = m,column = 2).value
                                        #r1 = sheetgl.cell(row = m,column = 3).value
                                   

                                    #if sn == 0 and co ==1:
                                        #print "k sn0"
                                        #print k
                                        #print ki
                                        

                                    if unicode(kil) in unicode(k):

                                    #if unicode(kil)== unicode(k):
                                       # print k, m, co, sheetgl
                                        #print "True"
                                        print ("Found '%s' in file %s,"%(k, ptf))
                                        print ("%s, col. %s, row %s"%(sheetv, co, m))
                                        print ("Content of row %s: "%(m))

                                        for cellrow in range(int(firstcol), int(col_countv)):
                                            printcell = sheetv.cell(row = m,column = cellrow).value
                                            printcell = unicode(printcell)
                                            print ("Col. '%s': %s "%(cellrow, printcell))
                                        
                                        #print k, m, co, sheetv
                                        #if unicode(ki) in unicode(n1):
                                          #  print n1, m, sheetgl
                                       # if unicode(ki) in unicode(r1):
                                          #  print r1, m, sheetgl

                                        nf = nf + 1


                                        qi = raw_input('Edit row y/n? ')



                                        if qi=="y":

                                            editcol = raw_input('Enter Number of Edit-Column, e.g. for first column "1": ')            
                                            editcol = int(editcol)

                                            edtext = raw_input('Enter text for edit: ')
                                            edtext = unicode(edtext)
                                            



                                          

                                            if colFill != "None":

                                                sheetv.cell(row=m, column=editcol).fill = colFill

                                            mf = sheetv.cell(row=m, column=editcol).value

                                            if mf != None:
                                                print "Content in cell found:"
                                                print mf
                                                ea = raw_input("Edit anyway, y/n? ")
                                                if ea=="y":
                                                    sheetv.cell(row=m, column=editcol).value = edtext
                                                    print "Cell updated."
                                                    upd = "Yes"
                                                if ea=="n":
                                                    mf = sheetv.cell(row=m, column=markcol).value
                                                    print "Cell not updated."
                                                    upd = "No"

                                            if mf == None:
                                                sheetv.cell(row=m, column=editcol).value = edtext
                                                print "Cell updated."
                                                upd = "Yes"
                                                

                                            

                                            if comcol != "None":
                                                cf = sheetv.cell(row=m, column=comcol).value
                                                if cf != None:
                                                    print "Content in comment cell found:"
                                                    print cf
                                                    eac = raw_input("Edit anyway, y/n? ")
                                                    if eac=="y":
                                                        comtext = raw_input('Enter comment: ')
                                                        comtext = unicode(comtext)
                                                        sheetv.cell(row=m, column=comcol).value = comtext
                                                        print "Comment cell updated."
                                                        upd = "Yes"
                                                    if eac=="n":
                                                        
                                                        print "Comment cell not updated."                                                   
                                                        upd = "No"

                                                if cf == None:
                                                    comtext = raw_input('Enter comment: ')
                                                    comtext = unicode(comtext)
                                                    sheetv.cell(row=m, column=comcol).value = comtext
                                                    print "Comment cell updated."
                                                    upd = "Yes"

                                                
                                              
                                            
                                            bookkey.save(filename = ptf)
                                            

                                        if qi=="n":
                                            upd = "No"
                                            print "File not updated."                    
  


                                        nu=nu+1

                                        wsgi1.cell(row=1, column=1).value = 'Input'
                                        wsgi1.cell(row=1, column=2).value = 'Found item'
                                        wsgi1.cell(row=1, column=3).value = 'File'
                                        wsgi1.cell(row=1, column=4).value = 'Sheet'
                                        wsgi1.cell(row=1, column=5).value = 'Col.'
                                        wsgi1.cell(row=1, column=6).value = 'Row'
                                        wsgi1.cell(row=1, column=7).value = 'File updated?'
                                        
                                        

                                        wsgi1.cell(row=nu, column=1).value = ki
                                        wsgi1.cell(row=nu, column=2).value = k
                                        wsgi1.cell(row=nu, column=3).value = ptf
                                        wsgi1.cell(row=nu, column=4).value = unicode(sheetv)
                                        wsgi1.cell(row=nu, column=5).value = co
                                        wsgi1.cell(row=nu, column=6).value = m
                                        wsgi1.cell(row=nu, column=7).value = upd
                                        

                                        wbgi.save(filename = filenameGI)

                if nf < 2:
                    print "Not found."
                                     





################se univ fuzzy end#########################
                    



    def Manual(self):

          

        master = Tk()

        w = Label(master, text="\n\n    Manual created. The Manual file ('ManualSAM.txt') is in the Search+Mark folder.     \n\n", bg="green")
        w.pack()
        
        with open("ManualSAM.txt",'w') as gcman:
            
            gcman.write("Search+Mark Manual")
            gcman.write("\n\n")
            gcman.write("CONTENT:\n\n")
            gcman.write("INTRODUCTION" + "\n")
            gcman.write("THE PATHS FILE" + "\n")
            gcman.write("Search+Mark Specific (Exact)" + "\n")
            gcman.write("Search+Mark Specific (Fuzzy)" + "\n")
            gcman.write("Search+Mark Universal (Exact)" + "\n")
            gcman.write("Search+Mark Universal (Fuzzy)" + "\n")
            gcman.write("Search+Edit Universal (Exact)" + "\n")
            gcman.write("Search+Edit Universal (Fuzzy)" + "\n")


            
            gcman.write("\n\n")
            gcman.write("INTRODUCTION" + "\n\n")
            gcman.write("Search+Mark is a tool to search for and mark entries in multiple .xlsx spreadsheet files. Different settings allow to adjust the sensitivity of the search engine as well as the location, cell color and content of the mark and comment input. Search+Mark allows to assign different selections of multiple .xlsx spreadsheets and to store this selections permanent. Search+Mark informs the user about all important data, without the need to open the .xlsx file. Speed up the Search+Mark progress through tons of excel files, clever and smart, use Search+Mark, save time, save nerves. Search+Mark contains the editor tool Search+Edit. With Search+Edit it is possible to search in multiple .xlsx spreadsheet files and to edit any cell, without opening the spreadsheet." + "\n\n")

            gcman.write("Search+Mark is written by A.D.Klumpp using Python and the Python library openpyxl including jdcal and et_xmlfile (see license texts below or in the folders of the libraries). Search+Mark is released under the terms of the GNU General Public License (See http://www.gnu.org/licenses/). Copyright (C) 2016 A.D.Klumpp. Search+Mark is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY. The full copyright notices and the full license texts shall be included in all copies or substantial portions of the Software." + "\n\n")
            gcman.write("Python is released under the Python Software Foundation License (see https://www.python.org/download/releases/2.7.6/license/). Openpyxl is released under MIT/Expat license (see https://openpyxl.readthedocs.org/en/latest/). Kivy is released under the MIT License (see https://github.com/kivy/kivy/blob/master/LICENSE). jdcal is released under BSD (see https://pypi.python.org/pypi/jdcal). et_xmlfile is released under MIT (Home-page: https://bitbucket.org/openpyxl/et_xmlfile)." + "\n\n")
            gcman.write("Please read the full license texts Online or in the LICENSES.txt document, which is inside the Search+Mark folder." + "\n\n")
            gcman.write("" + "\n\n")

            
            gcman.write("\n\n")
            gcman.write("THE INPUT FILE" + "\n\n")
            gcman.write("Format: .xlsx" + "\n\n")
            gcman.write("Structure:\n\n")

            gcman.write("Sheet 1:" + "\n")
            gcman.write("Column 1: Paths to excel files" + "\n")
            gcman.write("Use one line per path" + "\n\n")
            gcman.write("Example:" + "\n")
            gcman.write("/home/user/KeyExcelFolder/Keys1998.xlsx" + "\n\n")

            gcman.write("Don't input formatting data" + "\n")
            gcman.write
            gcman.write("Open the input file (open the START-Menu and select the analysis program, see below) and follwo the instructions." + "\n")
            gcman.write("Don't start another mode, while program is running." + "\n")
            gcman.write("Restart the program in order to change between modes. Close the program via X button." + "\n")
            gcman.write("Open with LibreOffice, OpenOffice or MS Excel the Report file, which is now in the same location as the Search+Mark starter file." + "\n")
            gcman.write("" + "\n")
            gcman.write("" + "\n")
             
        

            gcman.write("\n\n")
            gcman.write("Search+Mark Mode: Specific (Exact)." + "\n")
            gcman.write("Specific (Search in one specific column), Exact match (word boundaries sensitive, case insenesitive)." + "\n")

            gcman.write("\n\n")
            gcman.write("Search+Mark Mode: Specific (Fuzzy)" + "\n")
            gcman.write("Universal (Search in all columns), Exact match (word boundaries insensitive, case insenesitive)." + "\n")

            gcman.write("\n\n")
            gcman.write("Search+Mark Mode: Universal (Exact)" + "\n")
            gcman.write("Universal (Search in one specific column), Exact match (word boundaries sensitive, case insenesitive)." + "\n")

            gcman.write("\n\n")
            gcman.write("Search+Mark Mode: Universal (Fuzzy)" + "\n")
            gcman.write("Universal (Search in all columns), Exact match (word boundaries insensitive, case insenesitive)." + "\n")

            gcman.write("\n\n")
            gcman.write("Search+Edit Mode: Universal (Exact)" + "\n")
            gcman.write("Search mode: Universal (Search in one specific column), Exact match (word boundaries sensitive, case insenesitive). Edit Mode allows edits of any cell." + "\n")

            gcman.write("\n\n")
            gcman.write("Search+Edit Mode: Universal (Fuzzy)" + "\n")
            gcman.write("Search mode: Universal (Search in all columns), Exact match (word boundaries insensitive, case insenesitive). Edit Mode allows edits of any cell." + "\n")


            gcman.write("\n\n")
            gcman.write("\n\n")



            #string = StringProperty('')

            


        mainloop()









    def Legal(self):


        master = Tk()

        w = Label(master, text="\n\n    LICENSES text created. The LICENSES file ('LICENSESsam.txt') is in the Search+Mark folder.    \n\n", bg="green")
        w.pack()
        
        with open("LICENSESsam.txt",'w') as gpl:



            gpl.write("CONTENT:\n")
            gpl.write(GCLH + "\n")
            gpl.write(OPLH + "\n")
            gpl.write(JDCALH + "\n")
            gpl.write(ETXMLLH + "\n")
            #gpl.write(KVLH + "\n")
            gpl.write(GPLH + "\n")
            gpl.write(PLH + "\n")
            
            gpl.write("\n\n")
            gpl.write(GCLH + "\n")
            gpl.write(GCLT + "\n")
            
            gpl.write("\n\n")
            gpl.write(OPLH + "\n")
            gpl.write(OPLT + "\n")
            
            gpl.write("\n\n")
            gpl.write(JDCALH + "\n")
            gpl.write(JDCALT + "\n")

            gpl.write("\n\n")
            gpl.write(ETXMLLH + "\n")
            gpl.write(ETXMLLT + "\n")

            #gpl.write("\n\n")
            #gpl.write(KVLH + "\n")
            #gpl.write(KVLT + "\n")

            gpl.write("\n\n")
            gpl.write(GPLH + "\n")
            gpl.write(GPLT + "\n")           

            gpl.write("\n\n")
            gpl.write(PLH + "\n")
            gpl.write(PLT + "\n")
                
            gpl.write("\n\n")
        
  



     
        
        
            #string = StringProperty('')

            


        mainloop()




            
                    

def main():

  


    root = Tk()
    ex = GCMain(root)
    root.geometry("700x180+500+500")
    root.mainloop()  


if __name__ == '__main__':
    main()
